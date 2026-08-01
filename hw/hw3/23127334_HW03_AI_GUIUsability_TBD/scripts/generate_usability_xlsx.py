from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]


def convert(source_name, target_name, sheet_name):
    source = root / "task-2-usability" / source_name
    target = root / "task-2-usability" / target_name
    with source.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 20
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(target)
    print(target)


convert("participant-list.csv", "participant-list.xlsx", "Participants")
convert("sus-summary.csv", "sus-ueqs-summary.xlsx", "SUS Summary")
