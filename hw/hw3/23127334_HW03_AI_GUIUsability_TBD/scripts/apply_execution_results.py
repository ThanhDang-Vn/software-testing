from pathlib import Path
import csv
import json
import shutil

from openpyxl import load_workbook


root = Path(__file__).resolve().parents[1]
checklist_dir = root / "task-1-gui-checklist"
evidence_dir = checklist_dir / "failed-screenshots"
json_path = checklist_dir / "execution-results.json"
xlsx_path = checklist_dir / "gui-checklist-v1-reviewed.xlsx"
csv_path = checklist_dir / "gui-checklist-v1-reviewed.csv"

payload = json.loads(json_path.read_text(encoding="utf-8"))
results = payload["results"]

overrides = {
    "GUI-L-005": {
        "status": "Failed",
        "actual": "The light-blue “Quên mật khẩu?” link (#3B82F6 on white) is approximately 3.68:1, below WCAG AA 4.5:1 for normal text.",
        "notes": "Computed-color contrast measurement in Chromium.",
        "evidence": "failed-screenshots/GUI-L-005.png",
    },
    "GUI-L-025": {
        "status": "Failed",
        "actual": "No explicit inline error, error summary, aria-invalid, or aria-describedby association is rendered for either required field.",
        "notes": "DOM and accessibility-attribute inspection after empty submission.",
        "evidence": "failed-screenshots/GUI-L-025.png",
    },
    "GUI-L-026": {
        "status": "Failed",
        "actual": "The dynamically inserted credential error has no role=alert, aria-live, or equivalent live-region semantics.",
        "notes": "DOM inspection after failed credential submission.",
        "evidence": "failed-screenshots/GUI-L-026.png",
    },
    "GUI-O-025": {
        "status": "Failed",
        "actual": "The Order History table has no caption and exposes no dedicated accessible table name.",
        "notes": "DOM/table semantic inspection.",
        "evidence": "failed-screenshots/GUI-O-025.png",
    },
    "GUI-O-006": {
        "status": "Failed",
        "actual": "Dates render with browser-default US ordering such as 7/26/2026 instead of deterministic Vietnamese day/month/year formatting.",
        "notes": "Observed across all five seeded order rows in Chromium.",
        "evidence": "failed-screenshots/GUI-O-006.png",
    },
}
results.update(overrides)

copies = {
    "GUI-L-025.png": "GUI-L-011A.png",
    "GUI-L-026.png": "GUI-L-024A.png",
    "GUI-O-025.png": "GUI-O-019.png",
    "GUI-O-006.png": "GUI-O-005.png",
}
for target, source in copies.items():
    shutil.copyfile(evidence_dir / source, evidence_dir / target)

bug_map = {
    "GUI-L-001": "BUG-GUI-001", "GUI-L-002": "BUG-GUI-001",
    "GUI-L-005": "BUG-GUI-002", "GUI-L-006": "BUG-GUI-002",
    "GUI-L-007": "BUG-GUI-003", "GUI-L-008": "BUG-GUI-003", "GUI-L-020": "BUG-GUI-003",
    "GUI-L-009": "BUG-GUI-004", "GUI-L-010": "BUG-GUI-004",
    "GUI-L-011A": "BUG-GUI-004", "GUI-L-025": "BUG-GUI-004", "GUI-L-026": "BUG-GUI-004",
    "GUI-L-013": "BUG-GUI-005",
    "GUI-L-024A": "BUG-GUI-006",
    "GUI-O-005": "BUG-GUI-007", "GUI-O-006": "BUG-GUI-007",
    "GUI-O-010": "BUG-GUI-008", "GUI-O-011": "BUG-GUI-008", "GUI-O-027": "BUG-GUI-008",
    "GUI-O-012": "BUG-GUI-009", "GUI-O-013": "BUG-GUI-009",
    "GUI-O-016": "BUG-GUI-010", "GUI-O-022": "BUG-GUI-010",
    "GUI-O-019": "BUG-GUI-011", "GUI-O-025": "BUG-GUI-011",
    "GUI-O-020": "BUG-GUI-012", "GUI-O-021": "BUG-GUI-012", "GUI-O-028": "BUG-GUI-012",
}

with csv_path.open(encoding="utf-8-sig", newline="") as handle:
    reader = csv.DictReader(handle)
    columns = reader.fieldnames
    rows = list(reader)

for row in rows:
    result = results[row["ID"]]
    row["Status"] = result["status"]
    row["Actual Result"] = result["actual"]
    row["Notes"] = result.get("notes") or "Observed during automated Playwright execution; see linked failure screenshot."
    row["Evidence"] = result.get("evidence", "")
    row["Bug ID or GitHub Issue"] = bug_map.get(row["ID"], "")

with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=columns)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(xlsx_path)
sheet = workbook["V1 - Reviewed"]
headers = [cell.value for cell in sheet[1]]
by_id = {row["ID"]: row for row in rows}
for row_number in range(2, sheet.max_row + 1):
    item = by_id[sheet.cell(row_number, 1).value]
    for field in ["Status", "Actual Result", "Notes", "Evidence", "Bug ID or GitHub Issue"]:
        sheet.cell(row_number, headers.index(field) + 1).value = item[field]
workbook.save(xlsx_path)

payload["results"] = results
json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

counts = {
    status: sum(1 for item in rows if item["Status"] == status)
    for status in ["Passed", "Failed", "Blocked"]
}
print(counts)
