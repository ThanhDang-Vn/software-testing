from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]
checklist_dir = root / "task-1-gui-checklist"
target = checklist_dir / "gui-checklist-v1-reviewed.xlsx"


def read_csv(name):
    with (checklist_dir / name).open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


datasets = [
    ("V0 - Agent", read_csv("gui-checklist-v0-agent.csv")),
    ("V1 - Reviewed", read_csv("revised-gui-checklist.csv")),
    ("Non-GUI Supporting Tests", read_csv("non-gui-supporting-tests.csv")),
    ("Migration Log", read_csv("migration-log.csv")),
]

workbook = Workbook()
workbook.remove(workbook.active)
header_fill = PatternFill("solid", fgColor="1F4E78")

for title, rows in datasets:
    sheet = workbook.create_sheet(title)
    if not rows:
        continue
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[header] for header in headers])
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, 1):
        width = 18
        if header in {
            "Check",
            "Preconditions",
            "Test Steps",
            "Expected Result",
            "Why AI Missed It",
            "Human Review or Modification",
            "What was moved",
            "Why it was moved",
        }:
            width = 42
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

readme = workbook.create_sheet("README", 0)
readme.append(["Submission checklist workbook"])
readme.append(["V0 - Agent", "Original agent-generated baseline; do not execute."])
readme.append(["V1 - Reviewed", "Canonical GUI checklist to review and execute."])
readme.append(
    ["Non-GUI Supporting Tests", "Items retained outside the GUI assignment scope."]
)
readme.append(["Migration Log", "Traceability from original IDs to moved/split IDs."])
readme.append(
    [
        "Important",
        "Only V1 - Reviewed counts toward the GUI checklist total. "
        "Execution fields must remain blank until real execution.",
    ]
)
readme.column_dimensions["A"].width = 28
readme.column_dimensions["B"].width = 90
readme["A1"].font = Font(bold=True, size=14)

workbook.save(target)
print(target)
