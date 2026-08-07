from pathlib import Path
import csv
import json
import re

from openpyxl import load_workbook


root = Path(__file__).resolve().parents[1]
directory = root / "task-1-gui-checklist"
issues = json.loads((directory / "github-created-issues.json").read_text(encoding="utf-8"))
links = {issue["bug_id"]: issue["url"] for issue in issues}

csv_path = directory / "gui-checklist-v1-reviewed.csv"
with csv_path.open(encoding="utf-8-sig", newline="") as handle:
    reader = csv.DictReader(handle)
    headers = reader.fieldnames
    rows = list(reader)
for row in rows:
    bug_id = row["Bug ID or GitHub Issue"]
    if bug_id in links:
        row["Bug ID or GitHub Issue"] = f"{bug_id} — {links[bug_id]}"
with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)

xlsx_path = directory / "gui-checklist-v1-reviewed.xlsx"
workbook = load_workbook(xlsx_path)
sheet = workbook["V1 - Reviewed"]
xlsx_headers = [cell.value for cell in sheet[1]]
column = xlsx_headers.index("Bug ID or GitHub Issue") + 1
for row_number in range(2, sheet.max_row + 1):
    value = sheet.cell(row_number, column).value
    if value in links:
        sheet.cell(row_number, column).value = f"{value} — {links[value]}"
workbook.save(xlsx_path)

report_path = directory / "bug-report.md"
report = report_path.read_text(encoding="utf-8")
for bug_id, url in links.items():
    pattern = rf"(## {re.escape(bug_id)} .*?\| GitHub Issue \| )Pending( \|)"
    report, count = re.subn(pattern, rf"\g<1>{url}\g<2>", report, count=1, flags=re.DOTALL)
    if count != 1:
        raise RuntimeError(f"Could not update {bug_id} in bug report")
report_path.write_text(report, encoding="utf-8")

links_path = directory / "github-issues-links.md"
lines = [
    "# GitHub Issue Links",
    "",
    "- Repository: https://github.com/ThanhDang-Vn/software-testing",
    "- Labels: none requested",
    "",
    "| Bug ID | GitHub Issue |",
    "| --- | --- |",
]
for bug_id in sorted(links):
    lines.append(f"| {bug_id} | {links[bug_id]} |")
links_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
print(f"Updated {len(links)} issue links")
