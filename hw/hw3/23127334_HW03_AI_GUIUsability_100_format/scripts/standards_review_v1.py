from pathlib import Path
from collections import Counter
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DIR = ROOT / "task-1-gui-checklist"
V1 = DIR / "gui-checklist-v1-reviewed.xlsx"
FINAL_CSV = DIR / "gui-checklist-v1-reviewed.csv"

WCAG = "https://www.w3.org/WAI/WCAG22/Understanding/"
WAI_FORMS = "https://www.w3.org/WAI/tutorials/forms/"
WAI_LABELS = "https://www.w3.org/WAI/tutorials/forms/labels/"
WAI_VALIDATION = "https://www.w3.org/WAI/tutorials/forms/validation/"
WAI_NOTIFY = "https://www.w3.org/WAI/tutorials/forms/notifications/"
APG_ALERT = "https://www.w3.org/WAI/ARIA/apg/patterns/alert/"
APG_DIALOG = "https://www.w3.org/WAI/ARIA/apg/patterns/dialog-modal/"
APG_TABLE = "https://www.w3.org/WAI/ARIA/apg/patterns/table/"
GOV_INPUT = "https://design-system.service.gov.uk/components/text-input/"
GOV_ERROR = "https://design-system.service.gov.uk/components/error-message/"
GOV_TABLE = "https://design-system.service.gov.uk/components/table/"
GOV_BANNER = "https://design-system.service.gov.uk/components/notification-banner/"
NIELSEN = "https://www.nngroup.com/articles/ten-usability-heuristics/"

COLUMNS = [
    "ID", "Screen", "IA", "Category", "Check",
    "Requirement or Heuristic Source", "Preconditions", "Test Steps",
    "Expected Result", "Source", "Why AI Missed It",
    "Human Review or Modification", "Status", "Actual Result", "Notes",
    "Evidence", "Bug ID or GitHub Issue",
]
EXECUTION = ["Status", "Actual Result", "Notes", "Evidence", "Bug ID or GitHub Issue"]


def source_for(row):
    category = str(row["Category"]).lower()
    check = str(row["Check"]).lower()
    if "contrast" in category or "color" in category:
        return f"WCAG 2.2 SC 1.4.1, 1.4.3, 1.4.11 — {WCAG}"
    if any(word in category for word in ["zoom", "reflow", "scaling", "responsive"]):
        return f"WCAG 2.2 SC 1.4.4, 1.4.10 — {WCAG}"
    if "rtl" in category or "dark" in category:
        return "Exploratory — no explicit SUT requirement"
    if "focus" in category or "keyboard" in category:
        return f"WCAG 2.2 SC 2.1.1, 2.4.3, 2.4.7, 2.4.11 — {WCAG}"
    if "target" in category:
        return f"WCAG 2.2 SC 2.5.8 — {WCAG}"
    if "label" in category or "required" in category or "email type" in category:
        return f"WCAG 2.2 SC 1.3.1, 2.4.6, 3.3.2; WAI Forms — {WAI_LABELS}"
    if "password manager" in category:
        return f"WCAG 2.2 SC 1.3.5, 3.3.8; GOV.UK Text input — {GOV_INPUT}"
    if "error" in category or "empty submission" in category:
        return f"WCAG 2.2 SC 3.3.1, 3.3.3; WAI Forms; GOV.UK Error message — {WAI_NOTIFY} | {GOV_ERROR}"
    if "table" in category:
        return f"WCAG 2.2 SC 1.3.1; ARIA APG Table; GOV.UK Table — {APG_TABLE} | {GOV_TABLE}"
    if "cancel confirmation" in category:
        return f"Nielsen H3/H5; ARIA APG Dialog — {NIELSEN} | {APG_DIALOG}"
    if "loading" in category or "feedback" in category or "pending" in check:
        return f"Nielsen H1; WCAG 2.2 SC 4.1.3; ARIA APG Alert — {NIELSEN} | {APG_ALERT}"
    if "language" in category or "translation" in category or "date" in category or "currency" in category:
        return f"Nielsen H2/H4; WCAG 2.2 SC 3.1.1 — {NIELSEN} | {WCAG}"
    if "discoverability" in category or "navigation" in category:
        return f"Nielsen H4/H6; WCAG 2.2 SC 2.4.4 — {NIELSEN} | {WCAG}"
    return f"Nielsen H1/H4/H8; FR-21–FR-24 — {NIELSEN}"


book = load_workbook(V1)
sheet = book["V1 - Reviewed"]
headers = [cell.value for cell in sheet[1]]
rows = [
    dict(zip(headers, values))
    for values in sheet.iter_rows(min_row=2, values_only=True)
]

for row in rows:
    for field in COLUMNS:
        row.setdefault(field, "")
        if row[field] is None:
            row[field] = ""
    row["Requirement or Heuristic Source"] = source_for(row)
    expected = str(row["Expected Result"]).strip()
    if expected and not expected.upper().startswith("PASS IF"):
        row["Expected Result"] = f"PASS if {expected[0].lower() + expected[1:]}; otherwise FAIL"
    for field in EXECUTION:
        row[field] = ""
    if row["ID"] == "GUI-L-023":
        row["Category"] = "Exploratory — RTL"
        row["Requirement or Heuristic Source"] = "Exploratory — no explicit SUT requirement"
        row["Expected Result"] = "PASS if no text, control, or focus indicator overlaps or becomes unreachable; otherwise FAIL"


def new_row(id_, screen, ia, category, check, source, preconditions, steps, expected):
    row = {field: "" for field in COLUMNS}
    row.update({
        "ID": id_, "Screen": screen, "IA": ia, "Category": category,
        "Check": check, "Requirement or Heuristic Source": source,
        "Preconditions": preconditions, "Test Steps": steps,
        "Expected Result": expected, "Source": "Human Added",
        "Why AI Missed It": "[TO BE COMPLETED BY STUDENT]",
    })
    return row


rows.extend([
    new_row(
        "GUI-L-025", "Login", "IA-02", "Error association",
        "Each visible validation error is associated with its input",
        f"WCAG 2.2 SC 1.3.1, 3.3.1; WAI Forms — {WAI_NOTIFY}",
        "Login page open",
        "Submit each field with invalid or missing input; inspect the accessibility tree and move focus to the field",
        "PASS if each error is visible, names the affected field, and is programmatically associated with that field; otherwise FAIL",
    ),
    new_row(
        "GUI-L-026", "Login", "IA-04", "Dynamic error announcement",
        "A dynamically displayed login error is announced without moving focus unexpectedly",
        f"WCAG 2.2 SC 4.1.3; ARIA APG Alert — {APG_ALERT}",
        "Screen reader running; credentials produce a visible error",
        "Submit once and listen to the screen reader while observing keyboard focus",
        "PASS if the complete error is announced and focus remains at a logical control; otherwise FAIL",
    ),
    new_row(
        "GUI-L-027", "Login", "IA-02", "Focus not obscured",
        "Focused login controls are not hidden by other content",
        f"WCAG 2.2 SC 2.4.11 — {WCAG}",
        "Login page open at desktop and 320 px viewport",
        "Tab through every interactive control at both viewports",
        "PASS if every focused control and its focus indicator remain at least partially visible; otherwise FAIL",
    ),
    new_row(
        "GUI-L-028", "Login", "IA-01", "Text spacing",
        "Login content remains readable with WCAG text-spacing overrides",
        f"WCAG 2.2 SC 1.4.12 — {WCAG}",
        "Login page open",
        "Apply line-height 1.5, paragraph spacing 2em, letter spacing 0.12em, and word spacing 0.16em",
        "PASS if no label, link, input value, error, or button text is clipped or overlaps; otherwise FAIL",
    ),
    new_row(
        "GUI-L-029", "Login", "IA-01", "Orientation",
        "Login is usable in portrait and landscape orientation",
        f"WCAG 2.2 SC 1.3.4 — {WCAG}",
        "Mobile device or equivalent viewport",
        "Open the page in portrait, rotate to landscape, and complete all form interactions",
        "PASS if the same controls and content remain available and usable in both orientations; otherwise FAIL",
    ),
    new_row(
        "GUI-L-030", "Login", "IA-01", "Exploratory — dark mode",
        "Login remains readable when browser or OS dark mode is forced",
        "Exploratory — no explicit SUT requirement",
        "Browser or OS forced dark mode available",
        "Enable forced dark mode; inspect normal, focused, and error states",
        "PASS if all text, controls, focus indicators, and errors remain visible and distinguishable; otherwise FAIL",
    ),
    new_row(
        "GUI-O-025", "Order History", "IA-01", "Table caption",
        "Order table has a programmatically associated descriptive caption",
        f"WCAG 2.2 SC 1.3.1; GOV.UK Table — {GOV_TABLE}",
        "Order History contains at least one order",
        "Inspect the table accessibility tree and navigate to the table with a screen reader",
        "PASS if the table exposes a concise accessible caption identifying it as order history; otherwise FAIL",
    ),
    new_row(
        "GUI-O-026", "Order History", "IA-01", "Reading order",
        "Order information has a logical reading order at desktop and mobile widths",
        f"WCAG 2.2 SC 1.3.2 — {WCAG}",
        "Order History contains multiple rows",
        "Read the table with a screen reader at desktop width, then at 320 px width",
        "PASS if each row is announced in the visible column order without detached or repeated values; otherwise FAIL",
    ),
    new_row(
        "GUI-O-027", "Order History", "IA-02", "Dialog focus management",
        "Cancel confirmation keeps keyboard focus within the modal until it closes",
        f"ARIA APG Dialog — {APG_DIALOG}",
        "A visible cancel control opens a modal confirmation",
        "Open the dialog by keyboard; press Tab and Shift+Tab through controls; close with each available method",
        "PASS if focus enters the dialog, cycles within it, and returns to the invoking control after close; otherwise FAIL",
    ),
    new_row(
        "GUI-O-028", "Order History", "IA-01", "Localized table scrolling",
        "A wide order table does not cause page-wide horizontal scrolling",
        f"WCAG 2.2 SC 1.4.10; GOV.UK Table — {WCAG} | {GOV_TABLE}",
        "Order History populated; viewport 320 CSS px",
        "Scroll vertically and horizontally through the page and table region",
        "PASS if any required horizontal scrolling is confined to the table region and surrounding page content reflows; otherwise FAIL",
    ),
    new_row(
        "GUI-O-029", "Order History", "IA-04", "Status announcement",
        "A visible order-state change is announced to assistive technology",
        f"WCAG 2.2 SC 4.1.3; ARIA APG Alert — {WCAG} | {APG_ALERT}",
        "An action causes a visible order-state update",
        "Trigger the action once while a screen reader is running",
        "PASS if the updated state or confirmation is announced without requiring focus movement; otherwise FAIL",
    ),
])

ids = [row["ID"] for row in rows]
assert len(ids) == len(set(ids))
assert len(rows) > 40
assert {row["IA"] for row in rows} == {"IA-01", "IA-02", "IA-03", "IA-04"}

# Supporting tests retained outside the GUI checklist.
supporting = [
    ["GUI-L-011B", "GUI-L-011", "Empty form sends no login API request", "Integration"],
    ["GUI-L-012", "GUI-L-012", "Email whitespace normalization", "Functional"],
    ["GUI-L-015", "GUI-L-015", "Forgot-password route resolves correctly", "Functional"],
    ["GUI-L-016", "GUI-L-016", "Registration route resolves correctly", "Functional"],
    ["GUI-L-017", "GUI-L-017", "Successful authentication creates authenticated state", "Integration"],
    ["GUI-L-018", "GUI-L-018", "Authenticated-session revisit routing", "Integration"],
    ["GUI-L-019", "GUI-L-019", "Resistance to account enumeration", "Security"],
    ["GUI-L-024B", "GUI-L-024", "Repeated submit produces one login request", "Performance"],
    ["GUI-O-002B", "GUI-O-002", "Displayed orders match authenticated API data", "Integration"],
    ["GUI-O-014", "GUI-O-014", "Unauthenticated route authorization", "Security"],
    ["GUI-O-024B", "GUI-O-024", "Canceled state persists correctly", "Functional"],
]

changes = [
    ["All retained IDs", "Rewritten", "One objective, executable steps, and observable PASS/FAIL oracle."],
    ["GUI-L-011", "Split", "GUI-L-011A retains visible validation; GUI-L-011B supports network assertion."],
    ["GUI-L-024", "Split", "GUI-L-024A retains pending UI; GUI-L-024B supports request-count assertion."],
    ["GUI-O-002", "Split", "GUI-O-002A retains visible fields; GUI-O-002B supports API comparison."],
    ["GUI-O-024", "Split", "GUI-O-024A retains visible feedback; GUI-O-024B supports persistence."],
    ["GUI-L-023", "Marked Exploratory", "RTL has no explicit SUT requirement."],
    ["GUI-L-030", "Added Exploratory", "Dark mode has no explicit SUT requirement."],
    ["GUI-L-025–030; GUI-O-025–029", "Added", "Filled standards-backed accessibility and UI-state gaps."],
]

references = [
    ["WCAG 2.2 Understanding", WCAG],
    ["WAI Forms Tutorial", WAI_FORMS],
    ["WAI Form Labels", WAI_LABELS],
    ["WAI Form Validation", WAI_VALIDATION],
    ["WAI Form Notifications", WAI_NOTIFY],
    ["ARIA APG Alert", APG_ALERT],
    ["ARIA APG Dialog", APG_DIALOG],
    ["ARIA APG Table", APG_TABLE],
    ["GOV.UK Text Input", GOV_INPUT],
    ["GOV.UK Error Message", GOV_ERROR],
    ["GOV.UK Table", GOV_TABLE],
    ["GOV.UK Notification Banner", GOV_BANNER],
    ["Nielsen Ten Usability Heuristics", NIELSEN],
]

duplicates = [
    ["GUI-L-004 / GUI-L-021", "Related, not duplicate", "320 px responsive layout versus 200% zoom/reflow."],
    ["GUI-L-013 / GUI-L-014 / GUI-L-027", "Related, not duplicate", "Focus order, focus visibility, and focus obscuration have distinct oracles."],
    ["GUI-O-019 / GUI-O-025 / GUI-O-026", "Related, not duplicate", "Header association, caption, and reading order are separate semantics."],
    ["GUI-O-020 / GUI-O-021 / GUI-O-028", "Related, not duplicate", "Mobile reflow, text scaling, and localized table scrolling differ."],
    ["Exact duplicates", "None", "No pair has the same objective, steps, and expected result."],
]

coverage = Counter(row["IA"] for row in rows)
screen_coverage = Counter(row["Screen"] for row in rows)
category_coverage = Counter(
    "Exploratory" if str(row["Category"]).startswith("Exploratory") else
    "Accessibility" if any(term in str(row["Category"]).lower() for term in [
        "contrast", "target", "label", "keyboard", "focus", "zoom", "rtl",
        "spacing", "orientation", "semantics", "scaling", "caption", "reading"
    ]) else "GUI / Usability / Visible state"
    for row in rows
)

with FINAL_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

out = Workbook()
out.remove(out.active)


def add_sheet(title, headers_, data):
    ws = out.create_sheet(title)
    ws.append(headers_)
    for item in data:
        if isinstance(item, dict):
            ws.append([item.get(header, "") for header in headers_])
        else:
            ws.append(item)
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_ in ws.iter_rows(min_row=2):
        for cell in row_:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, header in enumerate(headers_, 1):
        ws.column_dimensions[get_column_letter(index)].width = (
            48 if header in {"Check", "Requirement or Heuristic Source", "Preconditions",
                             "Test Steps", "Expected Result", "Reason", "URL"} else 20
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


add_sheet("V1 - Reviewed", COLUMNS, rows)
add_sheet(
    "Non-GUI Supporting Tests",
    ["ID", "Traceability to Original ID", "Check", "Recommended Test Suite"],
    supporting,
)
add_sheet("Change Log", ["Original ID or Range", "Change", "Reason"], changes)
add_sheet("Source References", ["Source", "URL"], references)
add_sheet("Duplicate Report", ["IDs", "Assessment", "Reason"], duplicates)
add_sheet(
    "Coverage Summary",
    ["Dimension", "Value", "Count"],
    [["IA", key, value] for key, value in sorted(coverage.items())]
    + [["Screen", key, value] for key, value in sorted(screen_coverage.items())]
    + [["Scope", key, value] for key, value in sorted(category_coverage.items())]
    + [["Total", "GUI checklist items", len(rows)]
       ,["Total", "Non-GUI supporting tests", len(supporting)]],
)
out.save(V1)
print({"gui": len(rows), "supporting": len(supporting), "ia": dict(coverage)})
