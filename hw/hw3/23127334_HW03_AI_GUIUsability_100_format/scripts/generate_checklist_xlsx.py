from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


root = Path(__file__).resolve().parents[1]
source = root / "task-1-gui-checklist" / "gui-checklist.csv"
output_dir = root / "task-1-gui-checklist"

with source.open(encoding="utf-8-sig", newline="") as handle:
    source_rows = list(csv.DictReader(handle))

columns = [
    "ID",
    "Screen",
    "IA",
    "Category",
    "Check",
    "Preconditions",
    "Test Action",
    "Expected Result",
    "Status",
    "Actual Result",
    "Notes",
    "Evidence",
]


def write_csv(path, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row[column] for column in columns})


def write_xlsx(path, rows, version_label):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GUI Checklist"
    sheet.append(columns)
    for row in rows:
        sheet.append([row[column] for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    widths = [14, 17, 9, 20, 42, 32, 40, 48, 12, 42, 36, 28]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    status_column = columns.index("Status") + 1
    status_letter = get_column_letter(status_column)
    status_validation = DataValidation(
        type="list", formula1='"Passed,Failed,Blocked"', allow_blank=True
    )
    sheet.add_data_validation(status_validation)
    status_validation.add(f"{status_letter}2:{status_letter}{sheet.max_row}")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32

    summary = workbook.create_sheet("Summary")
    summary.append(["Version", version_label])
    summary.append(["Metric", "Value"])
    summary.append(["Designed items", f"=COUNTA('GUI Checklist'!A2:A{sheet.max_row})"])
    summary.append(
        [
            "Executed items",
            f'=COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Passed")'
            f'+COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Failed")'
            f'+COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Blocked")',
        ]
    )
    summary.append(
        [
            "Passed",
            f'=COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Passed")',
        ]
    )
    summary.append(
        [
            "Failed",
            f'=COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Failed")',
        ]
    )
    summary.append(
        [
            "Blocked",
            f'=COUNTIF(\'GUI Checklist\'!{status_letter}2:{status_letter}{sheet.max_row},"Blocked")',
        ]
    )
    for cell in summary[2]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 22
    workbook.save(path)
    print(path)


v0_rows = [row for row in source_rows if row["Source"] == "AI Initial"]
v1_rows = source_rows

write_csv(output_dir / "gui-checklist-v0-agent.csv", v0_rows)
write_csv(output_dir / "gui-checklist-v1-reviewed.csv", v1_rows)
write_xlsx(
    output_dir / "gui-checklist-v0-agent.xlsx",
    v0_rows,
    "v0 — Agent generated",
)
write_xlsx(
    output_dir / "gui-checklist-v1-reviewed.xlsx",
    v1_rows,
    "v1 — Human reviewed",
)
