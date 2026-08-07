from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parents[1]
checklist_dir = root / "task-1-gui-checklist"
combined_path = checklist_dir / "gui-checklist-v1-reviewed.xlsx"
combined = load_workbook(combined_path)


def sheet_rows(sheet):
    return [list(row) for row in sheet.iter_rows(values_only=True)]


v0_rows = sheet_rows(combined["V0 - Agent"])
v1_rows = sheet_rows(combined["V1 - Reviewed"])

# These four AI Initial rows belong to the untouched baseline. They were
# removed only during later human scope review and therefore must remain in v0.
v0_rows.extend(
    [
        [
            "GUI-O-001", "Order History", "IA-01", "Heading",
            "Order History has a clear page/section heading",
            "Logged in; open /profile",
            "Inspect heading hierarchy",
            "Exactly one page h1 exists and order history is clearly identified",
            "", "", "", "",
        ],
        [
            "GUI-O-003", "Order History", "IA-01", "Ownership",
            "Only the signed-in user's orders are visible",
            "Two users with distinct orders",
            "Open history as each user and compare",
            "No other user's order appears in either history",
            "", "", "", "",
        ],
        [
            "GUI-O-009", "Order History", "IA-02", "Cancel eligibility",
            "Cancel action appears only for pending or confirmed orders",
            "Orders exist in all states",
            "Inspect action cell for each state",
            "Cancel is present only for pending and confirmed",
            "", "", "", "",
        ],
        [
            "GUI-O-017", "Order History", "IA-04", "Empty state",
            "No-order account sees a friendly illustrated empty state",
            "Logged-in user has zero orders",
            "Open /profile",
            "Icon/illustration and friendly message explain the empty state",
            "", "", "", "",
        ],
    ]
)


def save_workbook(path, title, rows, description):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = (
            42 if column in {5, 6, 7, 8, 10, 11} else 18
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    info = workbook.create_sheet("README")
    info.append(["Version", description])
    info.append(["Execution", "Not executed; all execution fields are blank."])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 80
    info["A1"].font = Font(bold=True)
    workbook.save(path)


save_workbook(
    checklist_dir / "gui-checklist-v0-agent.xlsx",
    "V0 - Agent",
    v0_rows,
    "Original agent-generated baseline before any human review (32 items).",
)
save_workbook(
    checklist_dir / "gui-checklist-v1-reviewed.xlsx",
    "V1 - Reviewed",
    v1_rows,
    "Current checklist after human scope review (34 GUI items).",
)

print("v0:", len(v0_rows) - 1)
print("v1:", len(v1_rows) - 1)
