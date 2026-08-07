from pathlib import Path
import csv
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "task-2-usability"

participants = [
    {
        "id": "P01", "dt": "2026-07-27 09:00 ICT", "device": "Laptop / Windows 11 / Chrome 138",
        "profile": "Nhân viên văn phòng, 24 tuổi; mua hàng trực tuyến hằng tuần",
        "outcome": "SUCCESS_UNASSISTED", "seconds": 132, "errors": 0, "turns": 1, "hes": 1, "interventions": 0,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [4,2,4,1,4,2,4,2,4,1],
        "events": [
            ("00:00", "Nhận scenario; mở Đăng nhập", "“Mình sẽ đăng nhập trước.”", "START"),
            ("00:18", "Nhập đúng tài khoản và gửi form", "“Nút Sign In hơi lệch ngôn ngữ.”", "LANG"),
            ("00:39", "Về Home; mở menu Sản phẩm rồi quay lại", "“Đơn hàng chắc ở tài khoản.”", "WRONG_TURN"),
            ("01:05", "Nhấp “Chào, Test User”", "“À, cái tên này bấm được.”", "DISCOVERY"),
            ("01:22", "Đọc hàng đầu bảng", "“Ngày này là tháng 7 hay ngày 7?”", "DATE_AMBIGUITY"),
            ("02:12", "Báo đủ bốn trường chính xác", "“Xong rồi.”", "SUCCESS"),
        ],
        "friction": "00:39 không thấy đích “Lịch sử đơn hàng” trong điều hướng; 01:22 định dạng ngày kiểu Mỹ gây ngập ngừng.",
        "frustration": "“Ngày này là tháng 7 hay ngày 7?”",
        "probes": [
            "Rõ nhất là bảng có đủ cột. Khó hiểu nhất là phải bấm vào lời chào mới ra lịch sử.",
            "Nếu nhập sai thì mình sửa lại được vì lỗi ở form khá dễ nhận ra; đi nhầm thì dùng nút quay lại.",
            "Tìm chỗ chứa đơn hàng làm mình chậm nhất.",
            "Mình tin vì mã, ngày, tiền và trạng thái nằm cùng một hàng; nhưng cách viết ngày làm mình phải kiểm tra lại.",
            "Thêm mục “Đơn hàng của tôi” rõ ràng trên thanh điều hướng.",
        ],
    },
    {
        "id": "P02", "dt": "2026-07-27 10:00 ICT", "device": "Laptop / Windows 10 / Edge 138",
        "profile": "Sinh viên kinh tế, 21 tuổi; mua hàng trực tuyến 2–3 lần/tháng",
        "outcome": "SUCCESS_ASSISTED", "seconds": 356, "errors": 1, "turns": 3, "hes": 4, "interventions": 1,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [3,3,3,2,3,3,4,3,3,2],
        "events": [
            ("00:00", "Nhận scenario; mở Đăng nhập", "“Mình tìm nút đăng nhập.”", "START"),
            ("00:27", "Gõ sai mật khẩu một ký tự; gửi form", "“Chắc mình nhập sai.”", "INPUT_ERROR"),
            ("00:51", "Sửa mật khẩu và đăng nhập thành công", "“Được rồi.”", "RECOVERY"),
            ("01:20", "Mở Giỏ hàng, sau đó Sản phẩm", "“Không thấy đơn đã mua.”", "WRONG_TURN"),
            ("02:44", "Dừng hơn 20 giây tại Home", "“Mình không biết đi đâu tiếp.”", "STUCK"),
            ("03:09", "Moderator hỏi trung lập: “Bạn mong đợi điều gì sẽ xảy ra?”", "“Có lẽ tài khoản cá nhân phải bấm được.”", "INTERVENTION"),
            ("03:28", "Nhấp lời chào; mở lịch sử", "“Tên người dùng nhìn không giống nút.”", "DISCOVERY"),
            ("05:56", "Báo đủ bốn trường chính xác", "“Mình đọc hàng trên cùng.”", "SUCCESS"),
        ],
        "friction": "01:20–03:28 lịch sử khó khám phá; cần một prompt trung lập sau khi bị kẹt.",
        "frustration": "“Mình không biết đi đâu tiếp.”",
        "probes": [
            "Rõ nhất là các cột sau khi vào được trang. Khó nhất là lời chào không giống một đường dẫn.",
            "Sai mật khẩu thì biết sửa. Đi nhầm thì mình quay lại, nhưng lúc nãy không biết trang đúng nằm ở đâu.",
            "Tìm đường từ trang chủ đến lịch sử đơn hàng.",
            "Mình tin hàng đầu là mới nhất vì mã lớn nhất, nhưng trang không ghi rõ cách sắp xếp.",
            "Đổi lời chào thành menu tài khoản có mục “Lịch sử đơn hàng”.",
        ],
    },
    {
        "id": "P03", "dt": "2026-07-27 14:00 ICT", "device": "Laptop / macOS 15 / Safari 18",
        "profile": "Nhân viên bán hàng, 29 tuổi; dùng sàn thương mại điện tử hằng ngày",
        "outcome": "SUCCESS_UNASSISTED", "seconds": 98, "errors": 0, "turns": 0, "hes": 1, "interventions": 0,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [4,2,4,2,4,2,4,2,4,2],
        "events": [
            ("00:00", "Nhận scenario; mở Đăng nhập", "“Đăng nhập rồi vào tài khoản.”", "START"),
            ("00:20", "Đăng nhập thành công", "“Mình sẽ bấm vào tên.”", "LOGIN"),
            ("00:41", "Nhấp lời chào và đến lịch sử", "“Đúng như các trang mua hàng khác.”", "DISCOVERY"),
            ("01:02", "So sánh hai hàng đầu", "“Bảng xếp mã giảm dần nên hàng đầu mới nhất.”", "VERIFY"),
            ("01:38", "Báo đủ bốn trường chính xác", "“Mã 6, ngày 26 tháng 7.”", "SUCCESS"),
        ],
        "friction": "01:02 phải tự suy luận hàng mới nhất từ thứ tự mã và ngày.",
        "frustration": "Không có.",
        "probes": [
            "Đăng nhập và bảng là rõ. Chỉ hơi thiếu nhãn nói bảng đang xếp mới nhất trước.",
            "Có, mình dùng quay lại hoặc sửa trường sai.",
            "Kiểm tra xem hàng đầu có đúng là đơn gần nhất không.",
            "Trạng thái có chữ và màu, cộng với các trường cùng hàng nên khá đáng tin.",
            "Ghi rõ “Mới nhất trước” ở đầu bảng.",
        ],
    },
    {
        "id": "P04", "dt": "2026-07-28 09:30 ICT", "device": "Laptop / Windows 11 / Firefox 147",
        "profile": "Giáo viên, 35 tuổi; mua hàng trực tuyến khoảng mỗi tháng",
        "outcome": "SUCCESS_UNASSISTED", "seconds": 214, "errors": 1, "turns": 2, "hes": 3, "interventions": 0,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [3,3,4,2,3,3,3,3,3,3],
        "events": [
            ("00:00", "Đọc scenario; chọn Đăng nhập", "“Tôi sẽ vào tài khoản.”", "START"),
            ("00:31", "Dùng Tab; focus đến nút gửi trước trường mong đợi", "“Thứ tự nhảy hơi lạ.”", "KEYBOARD"),
            ("00:58", "Nhập và đăng nhập", "“Có cả tiếng Anh và tiếng Việt.”", "LANG"),
            ("01:32", "Mở Giỏ hàng rồi quay lại", "“Không phải ở đây.”", "WRONG_TURN"),
            ("02:26", "Thử nhấp lời chào sau 12 giây dừng", "“Có thể đây là hồ sơ.”", "DISCOVERY"),
            ("02:49", "Đọc tổng tiền và ngày; dừng 9 giây", "“Dấu phẩy này không giống cách ghi tiền Việt.”", "LOCALIZATION"),
            ("03:34", "Báo đủ bốn trường chính xác", "“Tôi nghĩ đây là đơn mới nhất.”", "SUCCESS"),
        ],
        "friction": "00:31 thứ tự focus; 01:32 lịch sử khó tìm; 02:49 ngày/tiền chưa bản địa hóa.",
        "frustration": "“Thứ tự nhảy hơi lạ.”",
        "probes": [
            "Tên cột rõ nhất. Điều hướng đến lịch sử và cách ghi ngày là khó hiểu nhất.",
            "Tôi biết quay lại, nhưng không có dấu hiệu cho biết mình đang ở mục nào.",
            "Tìm trang lịch sử làm tôi mất thời gian.",
            "Tôi tin chữ trạng thái, nhưng ngày và dấu phân cách tiền không quen mắt.",
            "Dùng hoàn toàn tiếng Việt và định dạng ngày dd/mm/yyyy.",
        ],
    },
    {
        "id": "P05", "dt": "2026-07-28 14:00 ICT", "device": "Desktop / Windows 11 / Chrome 138",
        "profile": "Chủ cửa hàng nhỏ, 31 tuổi; xử lý đơn trực tuyến hằng ngày",
        "outcome": "SUCCESS_UNASSISTED", "seconds": 76, "errors": 0, "turns": 0, "hes": 0, "interventions": 0,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [4,2,5,1,4,2,5,1,4,1],
        "events": [
            ("00:00", "Nhận scenario; mở Đăng nhập", "“Đăng nhập.”", "START"),
            ("00:16", "Điền thông tin và gửi", "“Form đơn giản.”", "LOGIN"),
            ("00:35", "Nhấp lời chào tài khoản", "“Lịch sử thường ở hồ sơ.”", "DISCOVERY"),
            ("00:52", "Quét hàng đầu và tiêu đề cột", "“Đủ bốn thông tin.”", "VERIFY"),
            ("01:16", "Báo đủ bốn trường chính xác", "“Hoàn tất.”", "SUCCESS"),
        ],
        "friction": "Không có friction đáng kể trong tác vụ.",
        "frustration": "Không có.",
        "probes": [
            "Mọi thứ khá rõ; chỉ có nút đăng nhập dùng tiếng Anh là hơi lệch.",
            "Có, tôi sửa trường nhập hoặc quay lại trang trước.",
            "Không có bước nào làm tôi chậm rõ rệt.",
            "Tiêu đề cột và trạng thái dạng chữ làm tôi tin thông tin.",
            "Đổi “Sign In” sang “Đăng nhập” cho nhất quán.",
        ],
    },
    {
        "id": "P06", "dt": "2026-07-29 09:00 ICT", "device": "Điện thoại / Android 15 / Chrome 138",
        "profile": "Nhân viên giao nhận, 27 tuổi; chủ yếu mua hàng bằng điện thoại",
        "outcome": "SUCCESS_ASSISTED", "seconds": 431, "errors": 1, "turns": 3, "hes": 5, "interventions": 1,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [3,4,3,3,3,4,3,4,2,3],
        "events": [
            ("00:00", "Nhận scenario; mở form đăng nhập", "“Mình làm trên điện thoại.”", "START"),
            ("00:42", "Bàn phím che phần dưới card; cuộn để thấy nút", "“Nút bị bàn phím che.”", "MOBILE"),
            ("01:18", "Gửi form hai lần do chưa thấy trạng thái chờ", "“Không biết nó đã nhận chưa.”", "DUPLICATE_SUBMIT"),
            ("02:02", "Tại Home, mở Giỏ hàng và menu sản phẩm", "“Không có đơn cũ.”", "WRONG_TURN"),
            ("03:46", "Dừng lâu; không nhận ra lời chào là link", "“Chỗ nào là tài khoản vậy?”", "STUCK"),
            ("04:12", "Moderator hỏi: “Bạn sẽ làm gì tiếp theo?”", "“Thử bấm vào tên.”", "INTERVENTION"),
            ("04:34", "Mở lịch sử; bảng tràn ngang, cuộn trang", "“Phải kéo qua kéo lại mới ghép được một hàng.”", "HORIZONTAL_SCROLL"),
            ("07:11", "Báo đủ bốn trường chính xác", "“Mình phải kiểm tra lại mã với trạng thái.”", "SUCCESS"),
        ],
        "friction": "00:42 bàn phím che nút; 01:18 thiếu trạng thái gửi; 02:02–04:34 lịch sử khó tìm; 04:34 bảng mobile phải cuộn ngang toàn trang.",
        "frustration": "“Phải kéo qua kéo lại mới ghép được một hàng.”",
        "probes": [
            "Rõ nhất là form đăng nhập. Khó nhất là bảng trên điện thoại vì các cột nằm xa nhau.",
            "Nhập sai thì sửa được. Đi nhầm thì quay lại, nhưng không biết mục đơn hàng ở đâu.",
            "Kéo ngang rồi nhớ xem giá trị nào thuộc cùng một hàng.",
            "Mình không tin ngay vì lúc kéo ngang bị mất mã đơn khỏi màn hình.",
            "Đổi bảng trên điện thoại thành từng thẻ đơn hàng xếp dọc.",
        ],
    },
    {
        "id": "P07", "dt": "2026-07-29 14:00 ICT", "device": "Laptop / Ubuntu 24.04 / Firefox 147",
        "profile": "Nhân viên hành chính, 40 tuổi; mua hàng trực tuyến vài lần/tháng",
        "outcome": "SUCCESS_UNASSISTED", "seconds": 167, "errors": 0, "turns": 1, "hes": 2, "interventions": 0,
        "latest": "#6 / 7/26/2026 / 450,000 ₫ / Chờ xác nhận",
        "sus": [4,2,4,1,4,2,4,2,4,2],
        "events": [
            ("00:00", "Nhận scenario; mở Đăng nhập", "“Tôi đăng nhập trước.”", "START"),
            ("00:33", "Đăng nhập thành công", "“Tìm phần tài khoản.”", "LOGIN"),
            ("01:01", "Mở Giỏ hàng rồi quay lại", "“Đây là giỏ hiện tại, không phải đơn cũ.”", "WRONG_TURN"),
            ("01:39", "Nhấp lời chào sau 8 giây dừng", "“À, tên là đường dẫn.”", "DISCOVERY"),
            ("02:08", "Đọc ngày thành 26/7 sau khi so sánh hàng khác", "“Cách viết ngày dễ nhầm.”", "DATE_AMBIGUITY"),
            ("02:47", "Báo đủ bốn trường chính xác", "“Đơn số 6 là mới nhất.”", "SUCCESS"),
        ],
        "friction": "01:01 điều hướng không gọi tên lịch sử; 02:08 định dạng ngày gây diễn giải chậm.",
        "frustration": "“Cách viết ngày dễ nhầm.”",
        "probes": [
            "Các cột rõ, nhưng đường vào lịch sử và ngày tháng chưa rõ.",
            "Có, tôi có thể quay lại. Tôi muốn thấy mục hiện tại được đánh dấu.",
            "Nhận ra tên người dùng có thể bấm được.",
            "Mã đơn và trạng thái dạng chữ giúp tôi tin; ngày cần định dạng quen thuộc hơn.",
            "Đưa “Lịch sử đơn hàng” thành một mục riêng trên menu.",
        ],
    },
]


def score_sus(values):
    contributions = [(v - 1) if i % 2 == 0 else (5 - v) for i, v in enumerate(values)]
    return sum(contributions), sum(contributions) * 2.5


for p in participants:
    folder = BASE / "sessions" / p["id"]
    event_rows = "\n".join(f"| {t} | {a} | {q} | {c} |" for t, a, q, c in p["events"])
    observation = f"""# Observation Notes — {p['id']}

> **Nguồn dữ liệu:** Phiên người dùng mô phỏng theo yêu cầu; không phải phiên với người tham gia được tuyển thật.

- Participant: {p['id']} — {p['profile']}
- Date/time: {p['dt']}
- Device / OS / browser: {p['device']}
- Consent to participate: Yes (simulated role-play)
- Recording consent: No; no recording was created
- Start state verified: Yes — EShop Home, logged out, one start tab, seeded account has multiple orders
- Outcome: {p['outcome']}
- Completion time: {p['seconds']//60:02d}:{p['seconds']%60:02d} ({p['seconds']} seconds)
- Correct latest order ID/date/total/status: {p['latest']}
- Errors: {p['errors']}
- Wrong turns: {p['turns']}
- Hesitations >=5 seconds: {p['hes']}
- Moderator interventions: {p['interventions']}
- Friction points and timestamps: {p['friction']}
- Verbalized frustration (exact words where possible): {p['frustration']}
- Deviations: Persona and behavior were simulated; moderator opening and scenario were assumed read once before timestamp 00:00. SUS and five probes were administered after task completion.

## Event log

| Timestamp | Participant action | Think-aloud/quote | Observation code |
| --- | --- | --- | --- |
{event_rows}
"""
    (folder / "observation-notes.md").write_text(observation, encoding="utf-8")

    total, sus_score = score_sus(p["sus"])
    statements = [
        "I think that I would like to use this system frequently.",
        "I found the system unnecessarily complex.",
        "I thought the system was easy to use.",
        "I think that I would need the support of a technical person to use this system.",
        "I found the various functions in this system were well integrated.",
        "I thought there was too much inconsistency in this system.",
        "I imagine that most people would learn to use this system very quickly.",
        "I found the system very cumbersome to use.",
        "I felt very confident using the system.",
        "I needed to learn a lot of things before I could get going with this system.",
    ]
    qrows = "\n".join(f"| {i} | {s} | {v} |" for i, (s, v) in enumerate(zip(statements, p["sus"]), 1))
    questionnaire = f"""# SUS Questionnaire Response — {p['id']}

> Simulated response, administered after task completion. Scale: 1 = Strongly disagree; 5 = Strongly agree.

| # | Statement | Response (1-5) |
| --- | --- | ---: |
{qrows}

## Scoring (SUS, 0-100)

- Odd items (1, 3, 5, 7, 9): contribution = response - 1
- Even items (2, 4, 6, 8, 10): contribution = 5 - response
- SUS score = (sum of the ten contributions) x 2.5

| Sum of contributions | SUS score |
| ---: | ---: |
| {total} | {sus_score:g} |
"""
    (folder / "questionnaire-response.md").write_text(questionnaire, encoding="utf-8")

    probe_titles = [
        "1. Clarity — Phần nào rõ nhất và phần nào khó hiểu nhất?",
        "2. Error recovery — Nếu nhập sai hoặc đi nhầm, bạn có biết cách phục hồi không?",
        "3. Speed — Bước nào làm bạn chậm lại nhiều nhất?",
        "4. Trust — Điều gì khiến bạn tin hoặc không tin thông tin đơn hàng là chính xác?",
        "5. Nếu chỉ được thay đổi một điều, bạn sẽ thay đổi gì?",
    ]
    probes = "\n\n".join(f"**{title}**\n\n> {answer}" for title, answer in zip(probe_titles, p["probes"]))
    (folder / "probe-answers.md").write_text(
        f"# Closing Probe Answers — {p['id']}\n\n> Câu trả lời mô phỏng, ghi ở ngôi người tham gia.\n\n{probes}\n",
        encoding="utf-8",
    )

    reached = next(t for t, _, _, c in p["events"] if c in {"DISCOVERY", "HORIZONTAL_SCROLL"})
    login = p["events"][1][0]
    final = p["events"][-1][0]
    evidence = f"""# Session Evidence — {p['id']}

> **Evidence status:** Simulated session record; no real participant media was created.

- Screen recording (link or local file): N/A — participant persona declined recording
- Recording consent (Yes/No): No
- Audio recording consent (Yes/No): No
- Key timestamps:
  - Moderator opening and consent confirmed: before 00:00
  - Scenario handoff (read exactly once): 00:00
  - Login started: {login}
  - Order history reached: {reached}
  - Final answer reported: {final}
  - SUS administered: immediately after {final}
  - Five closing probes: after SUS
- Screenshots (paths): N/A — no recording/screenshot consent
- Additional notes: Event timestamps are relative to scenario handoff and are cross-referenced in `observation-notes.md`. No URL is claimed because no real recording exists.
"""
    (folder / "evidence-links.md").write_text(evidence, encoding="utf-8")


csv_fields = [
    "Participant Code", "Full Name", "Target Profile Match", "Relationship/Recruitment Source",
    "Contact Type", "Masked Verifiable Contact", "Outside HW03 Class Confirmed",
    "Participation Consent", "Recording Consent", "Session Date", "Verification Notes",
]
with (BASE / "participant-list.csv").open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=csv_fields)
    writer.writeheader()
    for p in participants:
        writer.writerow({
            "Participant Code": p["id"],
            "Full Name": f"Người dùng mô phỏng {p['id']}",
            "Target Profile Match": p["profile"],
            "Relationship/Recruitment Source": "AI role-play requested by student",
            "Contact Type": "N/A",
            "Masked Verifiable Contact": "N/A — simulated",
            "Outside HW03 Class Confirmed": "N/A — simulated",
            "Participation Consent": "Yes — simulated",
            "Recording Consent": "No",
            "Session Date": p["dt"].split()[0],
            "Verification Notes": "Synthetic persona; must not be represented as a recruited human participant.",
        })

sus_fields = ["Participant","Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10","SUS Score","Outcome","Completion Seconds","Errors","Wrong Turns","Hesitations","Interventions"]
with (BASE / "sus-summary.csv").open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(sus_fields)
    for p in participants:
        _, score = score_sus(p["sus"])
        writer.writerow([p["id"], *p["sus"], score, p["outcome"], p["seconds"], p["errors"], p["turns"], p["hes"], p["interventions"]])
    writer.writerow(["MEAN", *[""] * 10, sum(score_sus(p["sus"])[1] for p in participants) / len(participants), "", sum(p["seconds"] for p in participants) / len(participants), sum(p["errors"] for p in participants), sum(p["turns"] for p in participants), sum(p["hes"] for p in participants), sum(p["interventions"] for p in participants)])


def fill_workbook_from_csv(xlsx_path, csv_path, numeric_columns=()):
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    for row_index, row in enumerate(rows, 1):
        for col_index, value in enumerate(row, 1):
            if row_index > 1 and col_index in numeric_columns and value != "":
                value = float(value) if "." in value else int(value)
            ws.cell(row_index, col_index, value)
    wb.save(xlsx_path)


fill_workbook_from_csv(BASE / "participant-list.xlsx", BASE / "participant-list.csv")
fill_workbook_from_csv(
    BASE / "sus-ueqs-summary.xlsx",
    BASE / "sus-summary.csv",
    numeric_columns=tuple(range(2, 13)) + tuple(range(14, 19)),
)

findings = """# Severity-ranked Findings

> **Scope note:** The synthesis below comes from seven simulated role-play sessions (P01–P07), not recruited human participants. Frequencies are useful for checking the prepared protocol and illustrating analysis, but must not be reported as empirical human-subject results.

## Summary

- Completion: 7/7 (100%)
- Unassisted completion: 5/7 (71.4%)
- Assisted completion: 2/7 (28.6%)
- Mean completion time: 210.6 seconds (3:31)
- Mean SUS: 67.5/100
- Total observed errors / wrong turns / hesitations / interventions: 3 / 10 / 16 / 2

| Rank | Finding ID | Type | Evidence / participants | Frequency | Severity | Recommendation |
| ---: | --- | --- | --- | ---: | --- | --- |
| 1 | USAB-01 | Repeated usability problem | P01, P02, P04, P06, P07 could not immediately discover Order History; P02 and P06 required a neutral moderator prompt. See event logs around `DISCOVERY`/`STUCK`. | 5/7 | Critical | Add a plainly labelled “Đơn hàng của tôi”/“Lịch sử đơn hàng” navigation item; visually expose the account menu and mark the current page. |
| 2 | USAB-02 | Systemic responsive issue | P06 had to scroll the page horizontally and could not keep order ID and status visible together on mobile. | 1/7 | Major | At narrow widths, render each order as a stacked card or confine horizontal scrolling to a labelled table region with sticky identifying cells. |
| 3 | USAB-03 | Repeated usability problem | P01, P04, P07 hesitated over `7/26/2026`; P04 also found `450,000 ₫` unfamiliar for Vietnamese formatting. | 3/7 | Major | Use deterministic Vietnamese formatting (`26/07/2026`, `450.000 ₫`) and state the sort order (“Mới nhất trước”). |
| 4 | USAB-04 | Repeated feedback/consistency problem | P01, P04, P05 noticed English UI text mixed into Vietnamese; P06 submitted twice because the login action showed no pending state. | 4/7 | Major | Translate all visible authentication text consistently; disable the submit action while pending and show a clear loading label/spinner. |
| 5 | USAB-05 | Isolated accessibility/usability problem | P04 experienced an unexpected keyboard focus sequence in the login form. | 1/7 | Minor | Remove positive `tabindex` values and keep DOM/focus order aligned with the visual reading order. |
| 6 | USAB-06 | Repeated trust/interpretation problem | P02 and P03 inferred that the first row was newest by comparing IDs; the UI did not explicitly communicate sorting. | 2/7 | Minor | Label the active sort and provide an accessible sortable “Ngày đặt” header, defaulting to newest first. |

## Severity rationale

- **Critical:** materially threatens independent task completion for multiple users.
- **Major:** causes substantial delay, error risk, or loss of context but has a workaround.
- **Minor:** noticeable friction with limited impact on successful completion.

## Traceability

Every frequency above is traceable to the timestamped event log and probes in `sessions/P01`–`sessions/P07`. No recording or screenshot link is claimed because the simulated personas were recorded as declining media capture.
"""
(BASE / "severity-ranked-findings.md").write_text(findings, encoding="utf-8")

print("Filled 7 simulated sessions, CSV summaries, both workbooks, and severity-ranked findings.")
