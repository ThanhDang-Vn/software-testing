from pathlib import Path
import csv

from openpyxl import load_workbook


root = Path(__file__).resolve().parents[1]
directory = root / "task-1-gui-checklist"
csv_path = directory / "gui-checklist-v1-reviewed.csv"
xlsx_path = directory / "gui-checklist-v1-reviewed.xlsx"
fields = ["Status", "Actual Result", "Notes", "Evidence", "Bug ID or GitHub Issue"]

with csv_path.open(encoding="utf-8-sig", newline="") as handle:
    reader = csv.DictReader(handle)
    headers = reader.fieldnames
    rows = list(reader)
for row in rows:
    for field in fields:
        row[field] = ""
with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
    writer = csv.DictWriter(handle, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)

workbook = load_workbook(xlsx_path)
sheet = workbook["V1 - Reviewed"]
xlsx_headers = [cell.value for cell in sheet[1]]
for row_number in range(2, sheet.max_row + 1):
    for field in fields:
        sheet.cell(row_number, xlsx_headers.index(field) + 1).value = ""
workbook.save(xlsx_path)
print(f"Cleared execution fields for {len(rows)} V1 items")
