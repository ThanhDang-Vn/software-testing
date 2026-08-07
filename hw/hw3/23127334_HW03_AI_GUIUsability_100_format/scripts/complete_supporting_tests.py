from pathlib import Path

from openpyxl import load_workbook


root = Path(__file__).resolve().parents[1]
path = root / "task-1-gui-checklist" / "gui-checklist-v1-reviewed.xlsx"
workbook = load_workbook(path)
sheet = workbook["Non-GUI Supporting Tests"]
existing = {sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)}

rows = [
    [
        "GUI-O-003",
        "GUI-O-003",
        "Only the authenticated user's orders are returned and displayed",
        "Security",
    ],
    [
        "GUI-O-009",
        "GUI-O-009",
        "Cancel eligibility follows the order-state business rule",
        "Functional",
    ],
]
for row in rows:
    if row[0] not in existing:
        sheet.append(row)

# Keep execution/result fields out of this supporting table; it is a scope map.
workbook.save(path)
print(f"Supporting tests: {sheet.max_row - 1}")
