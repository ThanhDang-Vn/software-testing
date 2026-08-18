import json
import base64
import re
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "reports" / "newman"
MACHINE_REPORT_DIR = ROOT / ".tools" / "newman-results"
WORKBOOK = ROOT / "testcases" / "23127334_HW06_API_TestCases.xlsx"
SUMMARY = REPORT_DIR / "postman-run-summary.md"

RUNS = {
    "Register": ("register", "REG"),
    "Coupon": ("coupon", "CPN"),
    "Product": ("product", "PRD"),
}

# These runs did not establish the state/token required by the audited precondition.
SETUP_FAILURE = {
    "CPN-AI-003", "CPN-AI-007", "CPN-AI-021", "CPN-AI-022", "CPN-AI-023", "CPN-AI-024",
    "CPN-AI-030", "CPN-H-001", "CPN-H-004", "CPN-H-005", "CPN-H-006", "CPN-H-007",
    "PRD-AI-005",
}

# The generated request executed, but the full audited multi-action/iteration oracle was not automated.
SCRIPT_INCOMPLETE = {
    "REG-AI-010", "REG-AI-039", "REG-AI-040", "REG-H-001", "REG-H-003", "REG-H-006", "REG-H-007",
    "CPN-AI-020", "CPN-H-008", "CPN-H-009",
    "PRD-AI-018", "PRD-AI-037", "PRD-H-003", "PRD-H-004", "PRD-H-006", "PRD-H-009",
}


def parse_actual_log(path):
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    actual = {}
    compact = re.sub(r"\s*\u2502\s*", "", "".join(lines))
    for match in re.finditer(r"\[TC_ACTUAL_B64\]((?:REG|CPN|PRD)-(?:AI|H)-\d{3})\|(\d+)\|(\d{3})\|([A-Za-z0-9+/=]+)", compact):
        try:
            body = base64.b64decode(match.group(4)).decode("utf-8")
            actual.setdefault(match.group(1), []).append({"iteration": int(match.group(2)), "status": int(match.group(3)), "body": body})
        except (ValueError, UnicodeDecodeError):
            pass
    if actual:
        return actual
    i = 0
    while i < len(lines):
        if "'[TC_ACTUAL]'" not in lines[i]:
            i += 1
            continue
        block = []
        i += 1
        while i < len(lines):
            line = re.sub(r"^\s*\u2502 ?", "", lines[i])
            if line.strip() == "}":
                break
            block.append(line.strip())
            i += 1
        text = "".join(block)
        tc = re.search(r"TC_ID: '([^']+)'", text)
        status = re.search(r"status: (\d+)", text)
        body = re.search(r"body: '(.*)'$", text)
        if tc and status:
            actual.setdefault(tc.group(1), []).append({
                "iteration": 0,
                "status": int(status.group(1)),
                "body": body.group(1) if body else "<body log unavailable>",
            })
        i += 1
    return actual


def failure_map(report):
    grouped = defaultdict(list)
    for failure in report["run"]["failures"]:
        source = failure.get("source", {}).get("name", "")
        match = re.match(r"((?:REG|CPN|PRD)-(?:AI|H)-\d{3})", source)
        if not match:
            continue
        error = failure.get("error", {})
        message = re.sub(r"\s+", " ", str(error.get("message", error.get("name", "assertion failure")))).strip()
        assertion = str(error.get("test", "") or failure.get("at", "assertion"))
        grouped[match.group(1)].append(f"{assertion}: {message}" if assertion else message)
    return grouped


def cli_metadata(path):
    metadata = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        if line == "--- NEWMAN OUTPUT ---":
            break
        if "=" in line:
            key, value = line.split("=", 1)
            metadata[key] = value
    for line in reversed(path.read_text(encoding="utf-8-sig").splitlines()):
        if "=" in line:
            key, value = line.split("=", 1)
            if key in {"TIMESTAMP_END_UTC", "NEWMAN_EXIT_CODE"}:
                metadata[key] = value
        if line == "--- END NEWMAN OUTPUT ---":
            break
    return metadata


def bug_candidate(tc_id):
    if tc_id.startswith("REG"):
        if tc_id in {"REG-AI-019"}: return "BUG-REG-UNIQUENESS"
        if tc_id in {"REG-AI-036", "REG-AI-038", "REG-H-005"}: return "BUG-REG-ERROR-HANDLING"
        return "BUG-REG-VALIDATION"
    if tc_id.startswith("CPN"):
        if tc_id in {"CPN-AI-006", "CPN-AI-008", "CPN-AI-025", "CPN-AI-026", "CPN-AI-029", "CPN-AI-031", "CPN-H-002"}: return "BUG-CPN-AUTH-IDENTITY"
        if tc_id in {"CPN-AI-013"}: return "BUG-CPN-MIN-BOUNDARY"
        if tc_id in {"CPN-AI-038"}: return "BUG-CPN-ERROR-SCHEMA"
        if tc_id in {"CPN-AI-001", "CPN-AI-010", "CPN-AI-011", "CPN-AI-015", "CPN-AI-016", "CPN-AI-017", "CPN-AI-025", "CPN-AI-027", "CPN-AI-028", "CPN-AI-032", "CPN-AI-039", "CPN-H-010"}: return "BUG-CPN-CALCULATION"
        return "BUG-CPN-VALIDATION"
    if tc_id in {"PRD-AI-002", "PRD-AI-003", "PRD-AI-004", "PRD-AI-006", "PRD-H-001", "PRD-H-002"}: return "BUG-PRD-AUTHORIZATION"
    if tc_id in {"PRD-AI-039", "PRD-H-007"}: return "BUG-PRD-ERROR-HANDLING"
    return "BUG-PRD-VALIDATION"


def classify(tc_id, expected_status, failures):
    ambiguity = is_ambiguous(expected_status)
    if tc_id in SETUP_FAILURE:
        return "FAIL", "ENVIRONMENT/SETUP FAILURE", "Required state/token precondition was not established in this isolated folder run.", ""
    if tc_id in SCRIPT_INCOMPLETE:
        return "FAIL", "TEST SCRIPT BUG", "Primary request ran, but the audited chained/iteration/storage/UI oracle was not fully automated.", ""
    if failures:
        note = " Specification contains an ambiguity for this case; expected alternatives were not changed." if ambiguity else ""
        return "FAIL", "SUT BUG", "Response/assertion diverged from the audited specification." + note, bug_candidate(tc_id)
    if ambiguity:
        return "PASS", "SPEC AMBIGUITY", "Observed response is within the audited allowed alternatives; specification still needs one canonical contract.", ""
    return "PASS", "PASS", "All implemented assertions passed against the audited expected result.", ""


def is_ambiguous(expected_status):
    text = str(expected_status).lower()
    return "spec gap" in text or " or " in text or (" if " in text and len(re.findall(r"\b[1-5]\d\d\b", text)) > 1)


def main():
    wb = load_workbook(WORKBOOK)
    results = {}
    run_meta = {}
    cli_meta = {}
    ambiguity_records = []

    for sheet, (key, prefix) in RUNS.items():
        cli = REPORT_DIR / f"{key}-run.cli.txt"
        report_path = MACHINE_REPORT_DIR / f"{key}-run.json"
        actual = parse_actual_log(cli)
        report = json.loads(report_path.read_text(encoding="utf-8"))
        cli_meta[sheet] = cli_metadata(cli)
        failures = failure_map(report)
        run_meta[sheet] = report["run"]
        ws = wb[sheet]
        headers = {cell.value: cell.column for cell in ws[1]}
        counts = Counter()
        assert len(actual) == sum(1 for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]), (sheet, len(actual))
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row, headers["ID"]).value
            if not tc_id:
                continue
            observations = sorted(actual[tc_id], key=lambda x: x["iteration"])
            fail_messages = failures.get(tc_id, [])
            verdict, category, reason, bug_id = classify(tc_id, ws.cell(row, headers["expected status"]).value, fail_messages)
            expected_status_text = str(ws.cell(row, headers["expected status"]).value)
            if is_ambiguous(expected_status_text):
                ambiguity_records.append((tc_id, expected_status_text, "/".join(str(x["status"]) for x in observations), verdict, category))
            rendered = []
            for observed in observations:
                body = observed["body"]
                if len(body) > 500:
                    body = body[:500] + "…[truncated; full body in CLI evidence]"
                rendered.append(f"i{observed['iteration']}: HTTP {observed['status']}; body={body}")
            assertion_note = "; ".join(fail_messages[:3]) if fail_messages else "none"
            actual_text = f"{' | '.join(rendered)}; classification={category}; reasoning={reason}; failed assertions={assertion_note}"
            ws.cell(row, headers["actual result"]).value = actual_text
            ws.cell(row, headers["PASS/FAIL"]).value = verdict
            ws.cell(row, headers["bug ID"]).value = bug_id or None
            ws.cell(row, headers["evidence link"]).value = f"../reports/newman/{key}-run.cli.txt#{tc_id}; ../reports/newman/{key}-run.html"
            counts[category] += 1
        results[sheet] = counts

    summary_ws = wb["Summary"]
    execution_headers = ["Executed", "Execution PASS", "Execution FAIL", "SUT bug failures", "Test script bug failures", "Environment/setup failures", "Spec ambiguity pass"]
    existing_summary_headers = {summary_ws.cell(1, c).value: c for c in range(1, summary_ws.max_column + 1)}
    start_col = existing_summary_headers.get("Executed", summary_ws.max_column + 1)
    for offset, header in enumerate(execution_headers):
        col = start_col + offset
        summary_ws.cell(1, col).value = header
        summary_ws.cell(1, col)._style = copy(summary_ws.cell(1, start_col - 1)._style)
        summary_ws.cell(1, col).alignment = copy(summary_ws.cell(1, start_col - 1).alignment)
        summary_ws.column_dimensions[summary_ws.cell(1, col).column_letter].width = 22
    summary_rows = {summary_ws.cell(r, 1).value: r for r in range(2, summary_ws.max_row + 1)}
    for sheet in RUNS:
        c = results[sheet]
        values = [sum(c.values()), c["PASS"] + c["SPEC AMBIGUITY"], c["SUT BUG"] + c["TEST SCRIPT BUG"] + c["ENVIRONMENT/SETUP FAILURE"], c["SUT BUG"], c["TEST SCRIPT BUG"], c["ENVIRONMENT/SETUP FAILURE"], c["SPEC AMBIGUITY"]]
        row = summary_rows[sheet]
        for offset, value in enumerate(values):
            summary_ws.cell(row, start_col + offset).value = value
            summary_ws.cell(row, start_col + offset)._style = copy(summary_ws.cell(row, start_col - 1)._style)
            summary_ws.cell(row, start_col + offset).alignment = copy(summary_ws.cell(row, start_col - 1).alignment)
    total_values = [148, totals_pass := sum(c["PASS"] + c["SPEC AMBIGUITY"] for c in results.values()), 148 - totals_pass,
                    sum(c["SUT BUG"] for c in results.values()), sum(c["TEST SCRIPT BUG"] for c in results.values()),
                    sum(c["ENVIRONMENT/SETUP FAILURE"] for c in results.values()), sum(c["SPEC AMBIGUITY"] for c in results.values())]
    if "Total" in summary_rows:
        for offset, value in enumerate(total_values):
            summary_ws.cell(summary_rows["Total"], start_col + offset).value = value
            summary_ws.cell(summary_rows["Total"], start_col + offset)._style = copy(summary_ws.cell(summary_rows["Total"], start_col - 1)._style)
            summary_ws.cell(summary_rows["Total"], start_col + offset).alignment = copy(summary_ws.cell(summary_rows["Total"], start_col - 1).alignment)

    wb.save(WORKBOOK)

    lines = [
        "# Newman Postman Run Summary",
        "",
        f"Generated from isolated final runs at `{datetime.now(timezone.utc).isoformat()}`.",
        "",
        "## Execution strategy",
        "",
        "Each API was executed in a separate Newman invocation. Before every invocation the exact Node process listening on port 3000 was stopped, `server.js` was restarted, and `database.js` dropped/recreated/seeded SQLite. Each invocation ran `00 Setup`, one selected API folder, then `99 Verification-Teardown`.",
        "",
        "Expected results were not changed. A passing observed implementation behavior does not resolve an explicitly documented specification ambiguity.",
        "",
        "## Run results",
        "",
        "| API | Primary cases | PASS | SUT bug | Test script bug | Environment/setup failure | Iterations | Requests | Assertions passed | Assertions failed |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for sheet, (key, _) in RUNS.items():
        c = results[sheet]
        primary = sum(c.values())
        passed = c["PASS"] + c["SPEC AMBIGUITY"]
        stats = run_meta[sheet]["stats"]
        failed_assertions = stats["assertions"]["failed"]
        passed_assertions = stats["assertions"]["total"] - failed_assertions
        lines.append(f"| {sheet} | {primary} | {passed} | {c['SUT BUG']} | {c['TEST SCRIPT BUG']} | {c['ENVIRONMENT/SETUP FAILURE']} | {stats['iterations']['total']} | {stats['requests']['total']} | {passed_assertions} | {failed_assertions} |")
    total = sum(sum(c.values()) for c in results.values())
    totals = Counter()
    for c in results.values(): totals.update(c)
    total_requests = sum(run_meta[s]["stats"]["requests"]["total"] for s in RUNS)
    total_assertions = sum(run_meta[s]["stats"]["assertions"]["total"] for s in RUNS)
    total_failed_assertions = sum(run_meta[s]["stats"]["assertions"]["failed"] for s in RUNS)
    lines.append(f"| **Total** | **{total}** | **{totals['PASS'] + totals['SPEC AMBIGUITY']}** | **{totals['SUT BUG']}** | **{totals['TEST SCRIPT BUG']}** | **{totals['ENVIRONMENT/SETUP FAILURE']}** | **{sum(run_meta[s]['stats']['iterations']['total'] for s in RUNS)}** | **{total_requests}** | **{total_assertions - total_failed_assertions}** | **{total_failed_assertions}** |")

    lines += [
        "",
        "`PASS/FAIL` in the workbook is based on the full audited case, not merely the primary HTTP status. Therefore a request can have passing Newman assertions but be recorded FAIL when its required state was not established or its chained oracle was not automated.",
        "",
        "## Failure classification rules",
        "",
        "- **SUT BUG:** setup was valid and one or more implemented assertions contradicted the audited specification response, schema, business value or observable side effect.",
        "- **TEST SCRIPT BUG:** the request ran, but the collection did not execute the complete audited chained, multi-iteration, storage or UI verification. It is not reported as a SUT defect.",
        "- **ENVIRONMENT/SETUP FAILURE:** required disabled/usage-limit/concurrency state or expired token was not established. The response is retained, but no SUT conclusion is made.",
        "- **SPEC AMBIGUITY:** the audited expected status explicitly allows alternatives. A PASS means the observed result was among those alternatives; the oracle was not rewritten.",
        "",
        "## CLI reproducibility metadata",
        "",
        f"- Node: `{cli_meta['Register'].get('NODE_VERSION', 'unknown')}`",
        f"- Newman: `{cli_meta['Register'].get('NEWMAN_VERSION', 'unknown')}`",
        "- Reporter: `newman-reporter-htmlextra 1.23.1` with `--reporter-htmlextra-skipSensitiveData`.",
        "",
    ]
    for sheet, (key, _) in RUNS.items():
        meta = cli_meta[sheet]
        lines += [
            f"### {sheet}",
            "",
            f"- Timestamp start UTC: `{meta.get('TIMESTAMP_START_UTC', 'unknown')}`",
            f"- Timestamp end UTC: `{meta.get('TIMESTAMP_END_UTC', 'unknown')}`",
            f"- Exit code: `{meta.get('NEWMAN_EXIT_CODE', 'unknown')}` (non-zero because audited assertions failed).",
            f"- Data: `{meta.get('DATA_FILE', 'unknown')}`",
            "",
            "```powershell",
            meta.get("COMMAND", "<command unavailable>"),
            "```",
            "",
        ]
    lines += [
        "## Specification ambiguity observations",
        "",
        "These cases retain their audited alternative oracle. The primary classification remains SUT/setup/script when that issue prevents a conclusive ambiguity-only PASS.",
        "",
        "| TC_ID | Audited expected status | Actual status | Verdict | Primary classification |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for tc_id, expected, status, verdict, category in ambiguity_records:
        lines.append(f"| `{tc_id}` | {expected.replace('|', '/')} | {status} | {verdict} | {category} |")
    lines += [
        "",
        "## Evidence",
        "",
        "| API | CLI evidence (status/body/assertions) | HTML report | Backend reset logs |",
        "| --- | --- | --- | --- |",
        "| Register | [register-run.cli.txt](register-run.cli.txt) | [register-run.html](register-run.html) | [stdout](register-backend.stdout.log), [stderr](register-backend.stderr.log) |",
        "| Coupon | [coupon-run.cli.txt](coupon-run.cli.txt) | [coupon-run.html](coupon-run.html) | [stdout](coupon-backend.stdout.log), [stderr](coupon-backend.stderr.log) |",
        "| Product | [product-run.cli.txt](product-run.cli.txt) | [product-run.html](product-run.html) | [stdout](product-backend.stdout.log), [stderr](product-backend.stderr.log) |",
        "",
        "Every workbook row contains the actual HTTP status, actual response body (truncated only when necessary with a pointer to full CLI evidence), classification, failed assertion summary and evidence reference.",
        "",
        "## Important limitations discovered during execution",
        "",
        "- Coupon disabled/usage-limit/concurrency scenarios need dedicated setup fixtures or supporting state APIs before they can yield a valid SUT verdict.",
        "- Expired-JWT cases require a supplied expired token; no JWT was hard-coded or signed using the SUT secret.",
        "- Several human/state/security cases require multi-action or UI/storage verification beyond one primary request. These are explicitly classified as test-script defects rather than SUT bugs.",
        "- Newman JSON reporter records supporting `pm.sendRequest` executions under the parent item. The authoritative byte-accurate primary status/body is decoded from the `[TC_ACTUAL_B64]` entry in each CLI evidence file.",
        "- Machine JSON reports contain resolved runtime auth data by reporter design and are retained only under `.tools/newman-results/`, which is Git-ignored. They are not public submission artifacts.",
        "",
        "## Terminal screenshot instructions",
        "",
        "The CLI files above are direct, unedited Newman output captured by the runner. To display one in a real terminal for a screenshot:",
        "",
        "```powershell",
        "Get-Content -Raw hw\\hw6\\reports\\newman\\register-run.cli.txt",
        "```",
        "",
        "Capture the terminal window showing the `COMMAND`, timestamps, Node/Newman versions and the Newman totals/assertions table. Repeat for Coupon/Product if required. Alternatively rerun the exact command recorded above after starting a freshly seeded backend and capture its live output.",
        "",
        "Do not edit the CLI file, paste fabricated totals, crop different runs together, or modify the screenshot. If secrets appear, rerun with a clean/safe reporter configuration and take a new screenshot rather than redacting an existing image.",
        "",
    ]
    SUMMARY.write_text("\n".join(lines), encoding="utf-8")
    print(json.dumps({sheet: dict(counts) for sheet, counts in results.items()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
