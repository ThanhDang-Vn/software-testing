import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "testcases" / "23127334_HW06_API_TestCases.xlsx"
OUT = ROOT / "postman"
COLLECTION_PATH = OUT / "23127334_HW06_API_Testing.postman_collection.json"
LOCAL_ENV_PATH = OUT / "23127334_HW06_Local.postman_environment.json"
EXAMPLE_ENV_PATH = OUT / "23127334_HW06_Local.example.postman_environment.json"


def js(lines):
    return {"type": "text/javascript", "exec": lines.strip().splitlines()}


def event(kind, source):
    return {"listen": kind, "script": js(source)}


def expected_codes(value):
    codes = [int(x) for x in re.findall(r"\b[1-5]\d\d\b", str(value))]
    return list(dict.fromkeys(codes)) or [200]


def raw_request(name, method, path, body=None, auth="none", content_type="application/json",
                prerequest=None, tests=None, description=""):
    headers = []
    if content_type:
        headers.append({"key": "Content-Type", "value": content_type, "type": "text"})
    request = {
        "method": method,
        "header": headers,
        "url": {"raw": "{{baseUrl}}" + path, "host": ["{{baseUrl}}"], "path": path.strip("/").split("/")},
        "description": description,
    }
    if auth == "user":
        request["auth"] = {"type": "bearer", "bearer": [{"key": "token", "value": "{{userToken}}", "type": "string"}]}
    elif auth == "admin":
        request["auth"] = {"type": "bearer", "bearer": [{"key": "token", "value": "{{adminToken}}", "type": "string"}]}
    elif auth == "expired":
        request["auth"] = {"type": "bearer", "bearer": [{"key": "token", "value": "{{expiredToken}}", "type": "string"}]}
    elif auth == "malformed":
        request["auth"] = {"type": "bearer", "bearer": [{"key": "token", "value": "{{malformedToken}}", "type": "string"}]}
    elif auth == "wrong-scheme":
        headers.append({"key": "Authorization", "value": "Basic {{userToken}}", "type": "text"})
        request["auth"] = {"type": "noauth"}
    else:
        request["auth"] = {"type": "noauth"}
    if body is not None:
        request["body"] = {"mode": "raw", "raw": body,
                           "options": {"raw": {"language": "json" if content_type == "application/json" else "text"}}}
    events = []
    if prerequest:
        events.append(event("prerequest", prerequest))
    if tests:
        events.append(event("test", tests))
    item = {"name": name, "request": request}
    if events:
        item["event"] = events
    return item


def common_test_script(tc_id, api, codes, expected_schema, expected_side_effect):
    codes_json = json.dumps(codes)
    schema_text = json.dumps(str(expected_schema), ensure_ascii=False)
    side_text = json.dumps(str(expected_side_effect), ensure_ascii=False)
    endpoint = {"Register": "/api/admin/users", "Coupon": "/api/coupons", "Product": "/api/products"}[api]
    token = "adminToken" if api != "Coupon" else "userToken"
    success_keys = {"Register": ["message", "id"], "Coupon": ["success", "coupon_id", "discount_amount", "final_amount", "message"], "Product": ["message", "id"]}[api]
    return f"""
const TC_ID = {json.dumps(tc_id)};
const expectedStatuses = {codes_json};
const expectedSchemaText = {schema_text};
const expectedSideEffect = {side_text};
pm.collectionVariables.set('activeTestId', TC_ID);
const primaryStatus = pm.response.code;
const primarySucceeded = primaryStatus >= 200 && primaryStatus < 300;
console.log('[TC_ACTUAL]', {{ TC_ID, status: primaryStatus, body: pm.response.text() }});
const actualBodyBase64 = btoa(unescape(encodeURIComponent(pm.response.text())));
console.log(`[TC_ACTUAL_B64]${{TC_ID}}|${{pm.info.iteration}}|${{primaryStatus}}|${{actualBodyBase64}}`);
let submittedBody = null;
try {{ submittedBody = JSON.parse(pm.variables.replaceIn(pm.request.body && pm.request.body.raw || 'null')); }} catch (e) {{}}
const sideEffectIdentity = {"submittedBody && !Array.isArray(submittedBody) ? submittedBody.email : null" if api == "Register" else "submittedBody && !Array.isArray(submittedBody) ? submittedBody.name : null" if api == "Product" else "null"};

pm.test(`${{TC_ID}} | status follows audited specification`, function () {{
  pm.expect(expectedStatuses, `expected one of ${{expectedStatuses.join(', ')}}`).to.include(pm.response.code);
}});

pm.test(`${{TC_ID}} | content-type is application/json`, function () {{
  pm.expect(pm.response.headers.get('Content-Type') || '').to.match(/^application\\/json\\b/i);
}});

let body = null;
pm.test(`${{TC_ID}} | response is parseable JSON`, function () {{
  body = pm.response.json();
  pm.expect(body).to.be.an('object');
}});

if (body) {{
  const success = pm.response.code >= 200 && pm.response.code < 300;
  pm.test(`${{TC_ID}} | exact response schema`, function () {{
    const actualKeys = Object.keys(body).sort();
    const expectedKeys = success ? {json.dumps(success_keys)}.sort() : ['error'];
    pm.expect(actualKeys, expectedSchemaText).to.eql(expectedKeys);
    if (success) {{
      pm.expect(body.message).to.be.a('string');
      {"pm.expect(body.id).to.be.a('number').and.above(0);" if api in ('Register', 'Product') else "pm.expect(body.success).to.eql(true); pm.expect(body.coupon_id).to.be.a('number').and.above(0); pm.expect(body.discount_amount).to.be.a('number'); pm.expect(body.final_amount).to.be.a('number');"}
    }} else {{
      pm.expect(body.error).to.be.a('string');
    }}
    ['password','token','stack','sql','query'].forEach(k => pm.expect(body).not.to.have.property(k));
  }});

  {business_script(api)}
  {capture_script(api)}
}}

const verifyToken = pm.environment.get('{token}');
if (verifyToken) {{
  pm.sendRequest({{
    url: pm.environment.get('baseUrl') + '{endpoint}',
    method: 'GET',
    header: {{ 'Authorization': 'Bearer ' + verifyToken, 'X-Student-Id': pm.environment.get('studentId') }}
  }}, function (err, response) {{
    pm.test(`${{TC_ID}} | side-effect verification request succeeds`, function () {{
      pm.expect(err, expectedSideEffect).to.equal(null);
      pm.expect(response.code).to.eql(200);
    }});
    if (!err && response.code === 200) {{
      const rows = response.json();
      {side_effect_script(api)}
    }}
  }});
}}
"""


def business_script(api):
    if api == "Register":
        return """if (pm.response.code >= 200 && pm.response.code < 300) {
    pm.test(`${TC_ID} | registration business values`, () => pm.expect(body.message).to.eql('User registered successfully'));
  }"""
    if api == "Product":
        return """if (pm.response.code >= 200 && pm.response.code < 300) {
    pm.test(`${TC_ID} | product business values`, () => pm.expect(body.message).to.eql('Product created'));
  }"""
    return """if (pm.response.code >= 200 && pm.response.code < 300) {
    const sent = JSON.parse(pm.variables.replaceIn(pm.request.body.raw));
    const total = Number(sent.total_amount);
    const fixed = sent.code === 'BIGBUY' ? 50000 : sent.code === 'VIP100' ? 100000 : null;
    const discount = fixed === null ? total * 0.10 : fixed;
    pm.test(`${TC_ID} | coupon business calculation`, function () {
      pm.expect(body.discount_amount).to.eql(discount);
      pm.expect(body.final_amount).to.eql(total - discount);
    });
  }"""


def capture_script(api):
    if api == "Register":
        return """if (pm.response.code >= 200 && pm.response.code < 300 && body.id) {
    pm.environment.set('createdEmail', pm.variables.get('caseEmail'));
    const ids = JSON.parse(pm.collectionVariables.get('createdUserIds') || '[]'); ids.push(body.id);
    pm.collectionVariables.set('createdUserIds', JSON.stringify(ids));
  }"""
    if api == "Product":
        return """if (pm.response.code >= 200 && pm.response.code < 300 && body.id) {
    pm.environment.set('createdProductId', String(body.id));
    const ids = JSON.parse(pm.collectionVariables.get('createdProductIds') || '[]'); ids.push(body.id);
    pm.collectionVariables.set('createdProductIds', JSON.stringify(ids));
  }"""
    return ""


def side_effect_script(api):
    if api == "Register":
        return """const email = sideEffectIdentity;
      if (email && Array.isArray(rows) && !String(email).includes('test@eshop.com')) {
        const matches = rows.filter(x => String(x.email) === String(email));
        pm.test(`${TC_ID} | user persistence/absence`, () => pm.expect(matches.length).to.eql(primarySucceeded ? 1 : 0));
      }"""
    if api == "Product":
        return """const marker = sideEffectIdentity;
      if (marker && Array.isArray(rows)) {
        const matches = rows.filter(x => String(x.name) === String(marker));
        pm.test(`${TC_ID} | product persistence/absence`, () => pm.expect(matches.length).to.eql(primarySucceeded ? 1 : 0));
      }"""
    return """pm.test(`${TC_ID} | coupon catalog remains readable`, function () {
        pm.expect(rows).to.be.an('array');
        const code = pm.variables.get('caseCouponCode');
        if (['SAVE10','BIGBUY','VIP100','EXPIRED'].includes(code)) pm.expect(rows.some(x => x.code === code)).to.eql(true);
      });"""


def request_prerequest(tc_id, api, data_ids=None):
    data_ids = data_ids or []
    return f"""
pm.variables.set('TC_ID', {json.dumps(tc_id)});
pm.collectionVariables.set('activeTestId', {json.dumps(tc_id)});
const suffix = `${{pm.collectionVariables.get('runId')}}-i${{pm.info.iteration}}-${{{json.dumps(tc_id)}}}`.replace(/[^A-Za-z0-9-]/g, '-');
pm.variables.set('caseEmail', `hw06.${{{json.dumps(tc_id.lower())}}}.${{suffix}}@example.test`);
pm.variables.set('caseProductName', `HW06-${{{json.dumps(tc_id)}}}-${{suffix}}`);
pm.variables.set('caseCouponCode', 'SAVE10');
const dataDefaults = {{name:'HW06 User', password:'Valid123!', code:'SAVE10', total_amount:500000,
  user_id:Number(pm.environment.get('userId') || 2), name_prefix:`HW06-${{{json.dumps(tc_id)}}}`, price:100000,
  description:'HW06 product', imageUrl:'https://example.test/p.png',
  category_id:Number(pm.environment.get('categoryId') || 1), auth_profile:'admin'}};
Object.keys(dataDefaults).forEach(k => pm.variables.set('data_' + k, dataDefaults[k]));
const acceptedDataIds = {json.dumps(data_ids)};
const currentDataId = pm.iterationData.get('data_id');
if (currentDataId && acceptedDataIds.length && acceptedDataIds.includes(currentDataId)) {{
  Object.keys(pm.iterationData.toObject()).forEach(k => pm.variables.set('data_' + k, pm.iterationData.get(k)));
  pm.variables.set('activeDataId', currentDataId);
}} else {{
  pm.variables.unset('activeDataId');
}}
"""


def register_body(tc_id):
    obj = {"name": "HW06 User", "email": "{{caseEmail}}", "password": "Valid123!", "confirm_password": "Valid123!"}
    raw = None; ctype = "application/json"
    n = int(tc_id[-3:]) if "-AI-" in tc_id else None
    if n == 1:
        obj["name"] = "{{data_name}}"; obj["password"] = "{{data_password}}"; obj["confirm_password"] = "{{data_password}}"
    elif n == 2: obj["name"] = "A"
    elif n == 3: obj["name"] = "Nguyễn Văn Á"
    elif n == 4: obj.pop("name")
    elif n == 5: obj["name"] = None
    elif n == 6: obj["name"] = 123
    elif n == 7: obj["name"] = ""
    elif n == 8: obj["name"] = "   "
    elif n == 9: obj["name"] = "Robert'); DROP TABLE users;--"
    elif n == 10: obj["name"] = "<img src=x onerror=alert(1)>"
    elif n == 11: obj.pop("email")
    elif n == 12: obj["email"] = None
    elif n == 13: obj["email"] = 123
    elif n == 14: obj["email"] = ""
    elif n == 15: obj["email"] = "invalid.example.test"
    elif n == 16: obj["email"] = "@example.test"
    elif n == 17: obj["email"] = "local@"
    elif n == 18: obj["email"] = " {{caseEmail}} "
    elif n == 19: obj["email"] = "test@eshop.com"
    elif n == 20: obj["email"] = "x' OR '1'='1@example.test"
    elif n == 21: obj.pop("password")
    elif n == 22: obj["password"] = None
    elif n == 23: obj["password"] = 12345678
    elif n == 24: obj["password"] = "Va1!abc"
    elif n == 25: obj["password"] = obj["confirm_password"] = "Va1!abcd"
    elif n == 26: obj["password"] = obj["confirm_password"] = "valid123!"
    elif n == 27: obj["password"] = obj["confirm_password"] = "VALID123!"
    elif n == 28: obj["password"] = obj["confirm_password"] = "Validabc!"
    elif n == 29: obj["password"] = obj["confirm_password"] = "Valid1234"
    elif n == 30: obj["password"] = obj["confirm_password"] = "Valid123#"
    elif n == 31: obj["password"] = obj["confirm_password"] = "Valid123&"
    elif n == 32: obj.pop("confirm_password")
    elif n == 33: obj["confirm_password"] = None
    elif n == 34: obj["confirm_password"] = "Other123!"
    elif n == 35: obj = {}
    elif n == 36: raw = '{"name":"Broken","email":'
    elif n == 37: obj = []
    elif n == 38: ctype = "text/plain"
    elif n == 39: obj.update({"role": "admin", "id": 1, "is_admin": True})
    if tc_id == "REG-H-002": obj["email"] = "TEST@ESHOP.COM"
    elif tc_id == "REG-H-003": obj["email"] = " {{caseEmail}} "
    elif tc_id == "REG-H-004": obj["confirm_password"] = {"value": "Valid123!"}
    elif tc_id == "REG-H-005": ctype = "text/plain"
    elif tc_id == "REG-H-006": obj.update({"email": "test@eshop.com", "role": "admin", "is_admin": True, "permissions": ["*"]})
    elif tc_id == "REG-H-007": obj["name"] = "<img src=x onerror=window.__hw06_xss=1>"
    elif tc_id == "REG-H-008": raw = '{"name":"Duplicate","email":"{{caseEmail}}","email":"other.{{caseEmail}}","password":"Valid123!","password":"Other123!","confirm_password":"Valid123!"}'
    elif tc_id == "REG-H-009": obj.update({"__proto__": {"role": "admin"}, "constructor": {"prototype": {"role": "admin"}}, "user": {"role": "admin"}})
    return raw if raw is not None else json.dumps(obj, ensure_ascii=False), ctype


def coupon_body(tc_id):
    obj = {"code": "SAVE10", "total_amount": 500000, "user_id": "{{userId}}"}; raw = None
    n = int(tc_id[-3:]) if "-AI-" in tc_id else None
    if tc_id in ("CPN-AI-010", "CPN-AI-011", "CPN-AI-016", "CPN-AI-017"):
        obj = {"code": "{{data_code}}", "total_amount": "{{data_total_amount}}", "user_id": "{{data_user_id}}"}
    if n == 2: obj["code"] = "HW06_NOT_FOUND"
    elif n == 3: obj["code"] = "DISABLED_HW06"
    elif n in (4,19): obj["code"] = "EXPIRED"
    elif n in (5,9): obj["total_amount"] = 299999
    elif n == 10: obj["total_amount"] = 300000
    elif n == 11: obj["total_amount"] = 300001
    elif n == 12: obj.update(code="BIGBUY", total_amount=499999)
    elif n == 13: obj.update(code="BIGBUY", total_amount=500000)
    elif n == 14: obj.update(code="BIGBUY", total_amount=500001)
    elif n == 15: obj["total_amount"] = 500000
    elif n == 16: obj["total_amount"] = 300001
    elif n == 17: obj.update(code="BIGBUY", total_amount=600000)
    elif n == 18: obj.update(code="VIP100", total_amount=600000)
    elif n in (21,22): obj.update(code="VIP100", total_amount=500000)
    elif n == 23: obj.update(code="VIP100", total_amount=500000)
    elif n == 24: obj.update(code="VIP100", total_amount=500000)
    elif n == 25: obj.pop("user_id")
    elif n == 26: obj["user_id"] = 1
    elif n == 27: obj["user_id"] = "2"
    elif n == 28: obj["user_id"] = -1
    elif n == 33: obj.pop("code")
    elif n == 34: obj["code"] = ""
    elif n == 35: obj.pop("total_amount")
    elif n == 36: obj["total_amount"] = None
    elif n == 37: obj["total_amount"] = -1
    elif n == 38: raw = '{"code":"SAVE10","total_amount":'
    elif n == 39: obj.update({"discount_value": 100, "is_active": True, "max_uses_per_user": 999, "discount_amount": 500000})
    elif n == 40: obj = []
    if tc_id == "CPN-H-001": obj.pop("user_id", None)
    elif tc_id == "CPN-H-002": obj["user_id"] = 1
    elif tc_id == "CPN-H-003": obj["code"] = "SAVE10' OR '1'='1"
    elif tc_id in ("CPN-H-004", "CPN-H-005"): obj.update(code="VIP100", total_amount=500000)
    elif tc_id == "CPN-H-006": obj["total_amount"] = 1
    elif tc_id == "CPN-H-007": obj["code"] = "DISABLED_HW06"
    elif tc_id == "CPN-H-008": obj["code"] = {"value": "SAVE10"}
    elif tc_id == "CPN-H-009": obj["total_amount"] = {"value": 300000}
    elif tc_id == "CPN-H-010": raw = '{"code":"SAVE10","total_amount":299999,"total_amount":300001,"user_id":1,"user_id":{{userId}}}'
    if tc_id in ("CPN-AI-010", "CPN-AI-011", "CPN-AI-016", "CPN-AI-017"):
        obj = {"code": "{{data_code}}", "total_amount": "{{data_total_amount}}", "user_id": "{{data_user_id}}"}
    encoded = raw if raw is not None else json.dumps(obj, ensure_ascii=False)
    return encoded.replace('"{{userId}}"', '{{userId}}').replace('"{{data_total_amount}}"', '{{data_total_amount}}').replace('"{{data_user_id}}"', '{{data_user_id}}')


def product_body(tc_id):
    obj = {"name": "{{caseProductName}}", "price": 100000, "description": "HW06 product", "imageUrl": "https://example.test/p.png", "category_id": "{{categoryId}}"}; raw = None; ctype = "application/json"
    n = int(tc_id[-3:]) if "-AI-" in tc_id else None
    if tc_id in ("PRD-AI-001", "PRD-AI-002", "PRD-AI-003"):
        obj = {"name": "{{data_name_prefix}}-{{runId}}", "price": "{{data_price}}", "description": "{{data_description}}", "imageUrl": "{{data_imageUrl}}", "category_id": "{{data_category_id}}"}
    if n == 7: obj.pop("name")
    elif n == 8: obj["name"] = None
    elif n == 9: obj["name"] = 123
    elif n == 10: obj["name"] = ""
    elif n == 11: obj["name"] = "   "
    elif n == 12: obj["name"] = "A"
    elif n in (13,14,15): obj["name"] = "X" * {13:254,14:255,15:256}[n]
    elif n == 16: obj["name"] = "Sản phẩm thử nghiệm {{runId}}"
    elif n == 17: obj["name"] = "x'); DROP TABLE products;--{{runId}}"
    elif n == 18: obj["name"] = "<img src=x onerror=alert(1)>-{{runId}}"
    elif n == 19: obj.pop("price")
    elif n == 20: obj["price"] = None
    elif n == 21: obj["price"] = "100000"
    elif n == 22: obj["price"] = 0
    elif n == 23: obj["price"] = -1
    elif n == 24: obj["price"] = 0.01
    elif n == 26: obj["price"] = 9007199254740991
    elif n == 27: obj["price"] = True
    elif n == 28: obj.pop("category_id")
    elif n == 29: obj["category_id"] = None
    elif n == 30: obj["category_id"] = "1"
    elif n == 31: obj["category_id"] = 0
    elif n == 32: obj["category_id"] = 999999
    elif n == 33: obj["category_id"] = 2
    elif n == 34: obj.pop("description")
    elif n == 35: obj.pop("imageUrl")
    elif n == 36: obj.pop("description"); obj.pop("imageUrl")
    elif n == 37: obj.update({"id": 1, "owner_id": 1, "role": "admin", "created_at": "2000-01-01"})
    elif n == 38: obj = {}
    elif n == 39: raw = '{"name":"broken","price":'
    elif n == 40: obj = []
    if tc_id == "PRD-H-004": obj["imageUrl"] = "javascript:alert(1)"
    elif tc_id == "PRD-H-005": obj["category_id"] = 999999
    elif tc_id == "PRD-H-007": ctype = "text/plain"
    elif tc_id == "PRD-H-008": raw = '{"name":"{{caseProductName}}","price":100000,"price":-1,"category_id":1,"category_id":2}'
    elif tc_id == "PRD-H-009": obj["price"] = {"value": 100000}
    encoded = raw if raw is not None else json.dumps(obj, ensure_ascii=False)
    encoded = encoded.replace('"{{categoryId}}"', '{{categoryId}}').replace('"{{data_price}}"', '{{data_price}}').replace('"{{data_category_id}}"', '{{data_category_id}}')
    return encoded, ctype


def auth_for(tc_id, api):
    if api == "Register": return "none"
    if api == "Coupon":
        if tc_id in ("CPN-AI-006", "CPN-AI-008"): return "none"
        if tc_id == "CPN-AI-029": return "malformed"
        if tc_id == "CPN-AI-030": return "expired"
        if tc_id == "CPN-AI-031": return "wrong-scheme"
        if tc_id == "CPN-AI-032": return "admin"
        return "user"
    if tc_id in ("PRD-AI-002", "PRD-H-001"): return "none"
    if tc_id in ("PRD-AI-003", "PRD-AI-006", "PRD-H-002"): return "user"
    if tc_id == "PRD-AI-004": return "malformed"
    if tc_id == "PRD-AI-005": return "expired"
    return "admin"


def classify(technique):
    t = str(technique).lower()
    if any(x in t for x in ["auth", "security", "idor", "injection", "xss", "mass assignment", "role", "sec-"]): return "Security"
    if any(x in t for x in ["state", "lifecycle", "concurr", "race", "toctou", "replay", "referential"]): return "State"
    if any(x in t for x in ["schema", "malformed", "content", "media", "parser", "type"]): return "Schema"
    return "Domain"


def data_mapping(tc_id):
    return {
        "REG-AI-001": ["REG-DATA-VALID-UNIQUE"], "REG-AI-019": ["REG-DATA-DUPLICATE-SEED", "REG-DATA-REPLAY-CONTROL"],
        "CPN-AI-010": ["CPN-DATA-SAVE10-EQUAL-MIN"], "CPN-AI-011": ["CPN-DATA-SAVE10-ABOVE-MIN"],
        "CPN-AI-016": ["CPN-DATA-SAVE10-ABOVE-MIN"], "CPN-AI-017": ["CPN-DATA-BIGBUY-FIXED"],
        "PRD-AI-001": ["PRD-DATA-ADMIN-VALID"], "PRD-AI-002": ["PRD-DATA-GUEST-VALID"], "PRD-AI-003": ["PRD-DATA-USER-VALID"],
    }.get(tc_id, [])


def build_primary(row, api):
    tc_id = row["ID"]
    if api == "Register": body, ctype = register_body(tc_id); path = "/api/register"
    elif api == "Coupon": body, ctype, path = coupon_body(tc_id), "application/json", "/api/apply-coupon"
    else: body, ctype = product_body(tc_id); path = "/api/products"
    codes = expected_codes(row["expected status"])
    mapping = data_mapping(tc_id)
    desc = "\n".join([
        f"TC_ID: {tc_id}", f"Origin: {row['origin']}", f"Technique: {row['technique']}",
        f"Requirement/SEC: {row['requirement/SEC reference']}", f"Priority: {row['priority']}",
        f"Preconditions: {row['preconditions']}", f"Test data: {row['test data']}",
        f"Expected schema: {row['expected headers/schema/body']}", f"Expected side effect: {row['expected side effect']}",
        f"Cleanup: {row['cleanup']}", f"Data IDs: {', '.join(mapping) if mapping else 'request-specific/static partition'}",
    ])
    return raw_request(
        f"{tc_id} | {row['title']}", "POST", path, body, auth_for(tc_id, api), ctype,
        request_prerequest(tc_id, api, mapping),
        common_test_script(tc_id, api, codes, row["expected headers/schema/body"], row["expected side effect"]), desc)


def setup_items():
    login_test = lambda role: f"""pm.test('SETUP | {role} login status', () => pm.response.to.have.status(200));
const body=pm.response.json(); pm.expect(body.token).to.be.a('string').and.not.empty;
pm.environment.set('{role}Token', body.token);
if ('{role}' === 'user') pm.environment.set('userId', String(body.user.id));
pm.test('SETUP | {role} role', () => pm.expect(body.user.role).to.eql('{role}'));"""
    return [
        raw_request("SETUP-01 | Health check", "GET", "/api/products", tests="pm.test('SETUP | backend reachable', () => pm.response.to.have.status(200));"),
        raw_request("SETUP-02 | Login seeded user", "POST", "/api/login", json.dumps({"email":"{{userEmail}}","password":"{{userPassword}}"}), tests=login_test("user")),
        raw_request("SETUP-03 | Login seeded admin", "POST", "/api/login", json.dumps({"email":"{{adminEmail}}","password":"{{adminPassword}}"}), tests=login_test("admin")),
        raw_request("SETUP-04 | Resolve category", "GET", "/api/categories", tests="""pm.test('SETUP | categories available', () => pm.response.to.have.status(200));
const rows=pm.response.json(); pm.expect(rows).to.be.an('array').and.not.empty; pm.environment.set('categoryId', String(rows[0].id));"""),
    ]


def teardown_items():
    return [
        raw_request("VERIFY-01 | Verify captured product", "GET", "/api/products/{{createdProductId}}", auth="admin", tests="pm.test('VERIFY | product endpoint responds', () => pm.expect(pm.response.code).to.be.oneOf([200,404]));"),
        raw_request("CLEAN-01 | Delete captured product", "DELETE", "/api/products/{{createdProductId}}", auth="admin", tests="pm.test('CLEAN | product delete responds', () => pm.expect(pm.response.code).to.be.oneOf([200,404])); pm.environment.unset('createdProductId');"),
        raw_request("CLEAN-02 | Delete captured user", "DELETE", "/api/admin/users/{{createdUserId}}", auth="admin", tests="pm.test('CLEAN | user delete responds', () => pm.expect(pm.response.code).to.be.oneOf([200,404])); pm.collectionVariables.unset('createdUserId'); pm.environment.unset('createdEmail');"),
        raw_request("CLEAN-03 | Clear run secrets", "GET", "/api/products", tests="""pm.test('CLEAN | final health', () => pm.response.to.have.status(200));
['userToken','adminToken','expiredToken','createdProductId','createdEmail'].forEach(k => pm.environment.unset(k));
['createdUserIds','createdProductIds','activeTestId','activeDataId','runId'].forEach(k => pm.collectionVariables.unset(k));"""),
    ]


def collection_prerequest():
    return """const EXPECTED_STUDENT_ID = '23127334';
pm.environment.set('studentId', EXPECTED_STUDENT_ID);
const studentId = String(pm.environment.get('studentId') || '').trim();
pm.request.headers.upsert({ key: 'X-Student-Id', value: studentId });
const attachedHeader = pm.request.headers.get('X-Student-Id');
const timestamp = new Date().toISOString();
const resolvedUrl = pm.variables.replaceIn(pm.request.url.toString());
pm.test('Pre-request: X-Student-Id is attached and correct', function () {
  pm.expect(studentId).to.eql(EXPECTED_STUDENT_ID); pm.expect(attachedHeader).to.eql(EXPECTED_STUDENT_ID);
});
console.log('[HW06 REQUEST EVIDENCE]', { timestamp, method: pm.request.method, url: resolvedUrl, 'X-Student-Id': attachedHeader });
if (!pm.collectionVariables.get('runId')) pm.collectionVariables.set('runId', `${Date.now()}-${pm.variables.replaceIn('{{$randomUUID}}')}`);
if (!pm.collectionVariables.get('createdUserIds')) pm.collectionVariables.set('createdUserIds','[]');
if (!pm.collectionVariables.get('createdProductIds')) pm.collectionVariables.set('createdProductIds','[]');"""


def local_secret(env_name, key):
    if os.environ.get(env_name):
        return os.environ[env_name]
    if LOCAL_ENV_PATH.exists():
        try:
            current = json.loads(LOCAL_ENV_PATH.read_text(encoding="utf-8"))
            return next((x.get("value", "") for x in current.get("values", []) if x.get("key") == key), "")
        except (json.JSONDecodeError, OSError):
            return ""
    return ""


def make_environment(name, local):
    values = {
        "baseUrl": "http://localhost:3000", "studentId": "23127334",
        "userEmail": "test@eshop.com", "userPassword": local_secret("HW06_USER_PASSWORD", "userPassword") if local else "",
        "adminEmail": "admin@eshop.com", "adminPassword": local_secret("HW06_ADMIN_PASSWORD", "adminPassword") if local else "",
        "userToken": "", "adminToken": "", "expiredToken": "", "malformedToken": "not-a-jwt",
        "userId": "", "categoryId": "", "createdProductId": "", "createdEmail": "",
    }
    secret = {"userPassword", "adminPassword", "userToken", "adminToken", "expiredToken"}
    return {"id": name, "name": name, "values": [{"key": k, "value": v, "type": "secret" if k in secret else "default", "enabled": True} for k,v in values.items()], "_postman_variable_scope": "environment", "_postman_exported_using": "HW06 generator"}


def main():
    wb = load_workbook(WORKBOOK, data_only=True)
    api_folders = []
    for sheet, label in [("Register", "API1 Register"), ("Coupon", "API2 Coupon"), ("Product", "API3 Product")]:
        ws = wb[sheet]; headers = [c.value for c in ws[1]]
        groups = {x: [] for x in ["Domain", "State", "Security", "Schema"]}
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values[0]: continue
            row = dict(zip(headers, values)); groups[classify(row["technique"])].append(build_primary(row, sheet))
        api_folders.append({"name": label, "item": [{"name": k, "item": v} for k,v in groups.items()]})
    collection = {
        "info": {"_postman_id": "23127334-hw06-api-testing", "name": "23127334_HW06_API_Testing", "description": "Generated from the final audited HW06 workbook. Expected assertions follow specification, not observed defects.", "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"},
        "event": [event("prerequest", collection_prerequest())],
        "item": [{"name": "00 Setup", "item": setup_items()}, *api_folders, {"name": "99 Verification-Teardown", "item": teardown_items()}],
        "variable": [{"key": "runId", "value": ""}, {"key": "createdUserIds", "value": "[]"}, {"key": "createdProductIds", "value": "[]"}],
    }
    OUT.mkdir(parents=True, exist_ok=True)
    COLLECTION_PATH.write_text(json.dumps(collection, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    LOCAL_ENV_PATH.write_text(json.dumps(make_environment("23127334_HW06_Local", True), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    EXAMPLE_ENV_PATH.write_text(json.dumps(make_environment("23127334_HW06_Local_Example", False), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Generated {COLLECTION_PATH}")


if __name__ == "__main__":
    main()
