import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

def thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def col_w(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def hdr(ws, r, c, v, bg="1F4E79", fg="FFFFFF", wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin()
    return cell

def dat(ws, r, c, v, wrap=True, bold=False, bg=None, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = thin()
    return cell

C_EDGE = "FFF2CC"
C_PASS = "C6EFCE"
C_FAIL = "FFC7CE"
C_HEAD = "1F4E79"
C_SEC  = "2E74B5"
C_LROW = "EBF3FB"

TCS = [
    ("TC01","Bat Nguon","Xac minh thiet bi khoi dong dung cach","Da cam nguon dien","1. Cam dien cho thiet bi\n2. Nhan nut nguon","Man hinh sang len","Man hinh sang va chua hien thi nhiet do","Dat",False,""),
    ("TC02","Hien Thi Nhiet Do Mac Dinh","Xac minh nhiet do mac dinh khi khoi dong","Thiet bi da bat","1. Bat thiet bi\n2. Bam nut nhiet do","Man hinh hien thi 180C theo mac dinh","Man hinh hien thi 180C theo mac dinh","Dat",False,""),
    ("TC03","Tang Nhiet Do","Xac minh co the tang nhiet do","Nhan nut tang","1. Bat thiet bi\n2. Chon nut dieu chinh nhiet do\n3. Nhan tang nhieu lan\n4. Quan sat","Nhiet do tang 5C moi lan nhan, toi da 200C","Nhiet do tang 5C moi lan nhan","Dat",False,""),
    ("TC04","Giam Nhiet Do","Xac minh co the giam nhiet do","Nhan nut giam","1. Bat thiet bi\n2. Chon nut dieu chinh nhiet do\n3. Nhan giam nhieu lan\n4. Quan sat","Nhiet do giam 5C moi lan nhan, toi thieu 80C","Nhiet do giam 5C moi lan nhan","Dat",False,""),
    ("TC05","Cai Dat Bo Hen Gio","Xac minh co the cai dat thoi gian","Nhan nut hen gio, dieu chinh bang tang/giam","1. Bat thiet bi\n2. Nhan bieu tuong hen gio\n3. Dat 10 phut bang nut tang","Bo dem hien thi 10:00 va bat dau dem nguoc; hien thi thoi gian toi da de nguoi dung de nhan biet","Co hien thi bo dem nhung khong hien thi thoi gian toi da khien nguoi dung kho nhan biet","Khong dat",False,"BUG-01"),
    ("TC06 (Edge)","Tang thoi gian khi dang nau","Tang thoi gian nau khi cam thay chua du trong luc may van dang chay","Nhiet do 180C, thoi gian 1 phut, gio da lap, may dang chay","1. May dang hoat dong trong khoang nho hon 60s\n2. Nguoi dung bam tang thoi gian","Thoi gian se duoc tang len","Thiet bi bip coi canh bao khong cho tang thoi gian","Khong dat",True,"AI bo sot: AI tao TC dua tren luong tieu chuan, khong xet thay doi thong so giua chung\nBUG-02"),
    ("TC07","Tam Dung Va Tiep Tuc","Xac minh chuc nang tam dung/tiep tuc hoat dong dung","Thiet bi dang nau","1. Bat dau nau\n2. Nhan tam dung\n3. Nhan lai de tiep tuc","Qua trinh nau tam dung roi tiep tuc tu dung thoi diem truoc do","Qua trinh nau tam dung roi tiep tuc tu dung thoi diem truoc do","Dat",False,""),
    ("TC08","Thao Gio Khi Dang Nau","Xac minh tu dong tam dung khi keo gio ra","Thiet bi dang hoat dong","1. Bat dau nau\n2. Keo gio ra giua qua trinh\n3. Quan sat man hinh","Thiet bi tu dong tam dung de dam bao an toan","Thiet bi tu dong tam dung de dam bao an toan","Dat",False,""),
    ("TC09","Canh Bao Khi Hoan Thanh","Xac minh am bao phat ra khi het thoi gian","Dat thoi gian 1 phut","1. Dat thoi gian 1 phut\n2. Bat dau\n3. Cho hoan thanh","Thiet bi phat tieng bip va ngung gia nhiet","Thiet bi phat tieng bip va ngung gia nhiet","Dat",False,""),
    ("TC10","Tat Nguon Khi Dang Nau","Xac minh thiet bi dung an toan khi tat giua chung","Thiet bi dang nau o 180C","1. Bat dau nau\n2. Nhan nut nguon de tat","Thiet bi dung ngay lap tuc, man hinh tat","Thiet bi dung ngay lap tuc, man hinh tat","Dat",False,""),
    ("TC11","Gioi Han Nhiet Do Toi Da","Xac minh gioi han nhiet do toi da (200C)","Nhan tang lien tuc tu muc mac dinh","1. Bat thiet bi\n2. Tiep tuc nhan tang vuot qua 200C","Nhiet do dung o 200C; hien thi chi bao toi da de nguoi dung nhan biet","Chi phat tieng bip, khong co chi bao truc quan","Khong dat",False,"BUG-03"),
    ("TC12","Gioi Han Nhiet Do Toi Thieu","Xac minh gioi han nhiet do toi thieu (80C)","Nhan giam lien tuc tu muc mac dinh","1. Bat thiet bi\n2. Tiep tuc nhan giam xuong duoi 80C","Nhiet do dung o 80C; hien thi chi bao toi thieu de nguoi dung nhan biet","Chi phat tieng bip, khong co chi bao truc quan","Khong dat",False,"BUG-04"),
    ("TC13","Mat Dien Dot Ngot Khi Dang Nau","Xac minh hanh vi khi nguon dien bi ngat dot ngot","Thiet bi dang nau, rut dien dot ngot","1. Bat dau nau 180C/10 phut\n2. Rut dien dot ngot\n3. Cam lai dien\n4. Quan sat","Thiet bi khoi dong lai an toan va giu nguyen nhiet do/thoi gian truoc luc rut phich","Thiet bi khoi dong lai nhung mat het du lieu, phai cai dat lai","Khong dat",False,"BUG-05"),
    ("TC14","Nau Khi Gio Chua Duoc Lap Hoan Toan","Xac minh hanh vi an toan khi gio chua khoa hoan toan","Gio duoc lap chua khop hoan toan","1. Lap gio mot phan\n2. Nhan bat dau\n3. Quan sat","Thiet bi khong khoi dong HOAC hien thi canh bao","Khong chon nut khoi dong duoc; neu chon se nghe tieng bip canh bao","Dat",False,""),
    ("TC15","Nhan Nut Lien Tuc / Spam Dau Vao","Xac minh tinh on dinh khi cac nut duoc nhan lien tuc","Nhan lien tuc cac nut tang/giam nguon va hen gio","1. Bat thiet bi\n2. Nhan lien tuc tat ca cac nut 5-10 lan trong 2 giay\n3. Quan sat man hinh","Man hinh on dinh, khong treo hoac hien thi gia tri bat thuong","Man hinh on dinh, khong treo hoac hien thi gia tri bat thuong","Dat",False,""),
    ("TC16 (Edge)","Bam khoi dong khi chua cam dien","Kiem tra hanh vi khi nhan nut nguon ma chua cam dien","Thiet bi chua duoc cam vao o dien","1. Dam bao day nguon chua cam\n2. Nhan nut nguon\n3. Quan sat phan hoi","Thiet bi khong phan hoi, man hinh khong sang, khong co am thanh","Thiet bi khong phan hoi, man hinh khong sang, khong co am thanh","Dat",True,"AI bo sot: AI luon gia dinh thiet bi da cam dien san nhu precondition ngam dinh"),
    ("TC17 (Edge)","Bat/tat den chieu sang cua noi","Kiem tra chuc nang bat/tat den ben trong noi","Thiet bi da cam dien, nhan nut den","1. Cam dien va bat nguon\n2. Nhan nut den de bat\n3. Quan sat den ben trong noi\n4. Nhan lai de tat\n5. Quan sat den tat","Den bat khi nhan lan 1, tat khi nhan lan 2; trang thai den thay doi chinh xac","Den bat khi nhan lan 1, tat khi nhan lan 2; trang thai den thay doi chinh xac","Dat",True,"AI bo sot: AI khong biet tinh nang den rieng cua model Philips HD9252"),
]

EXEC = [
    ("TC01","Bat Nguon","Normal","Yes","Yes","No","Dat"),
    ("TC02","Hien Thi Nhiet Do Mac Dinh","Normal","Yes","No","No","Dat"),
    ("TC03","Tang Nhiet Do","Normal","Yes","No","No","Dat"),
    ("TC04","Giam Nhiet Do","Normal","Yes","No","No","Dat"),
    ("TC05","Cai Dat Bo Hen Gio","Normal","Yes","Yes","Yes","Khong dat"),
    ("TC06 (Edge)","Tang thoi gian khi dang nau","Edge Case","Yes","Yes","Yes","Khong dat"),
    ("TC07","Tam Dung Va Tiep Tuc","Normal","Yes","No","No","Dat"),
    ("TC08","Thao Gio Khi Dang Nau","Normal","Yes","No","No","Dat"),
    ("TC09","Canh Bao Khi Hoan Thanh","Normal","Yes","No","No","Dat"),
    ("TC10","Tat Nguon Khi Dang Nau","Normal","Yes","No","No","Dat"),
    ("TC11","Gioi Han Nhiet Do Toi Da","BVA","Yes","No","Yes","Khong dat"),
    ("TC12","Gioi Han Nhiet Do Toi Thieu","BVA","Yes","No","Yes","Khong dat"),
    ("TC13","Mat Dien Dot Ngot Khi Dang Nau","Normal","Yes","No","Yes","Khong dat"),
    ("TC14","Nau Khi Gio Chua Duoc Lap Hoan Toan","Normal","Yes","Yes","No","Dat"),
    ("TC15","Nhan Nut Lien Tuc / Spam Dau Vao","Normal","Yes","Yes","No","Dat"),
    ("TC16 (Edge)","Bam khoi dong khi chua cam dien","Edge Case","Yes","No","No","Dat"),
    ("TC17 (Edge)","Bat/tat den chieu sang cua noi","Edge Case","Yes","No","No","Dat"),
]

BUGS = [
    ("BUG-01","TC05","Khong hien thi gioi han thoi gian toi da khi cai hen gio","Thap","Man hinh khong thay doi, chi phat tieng bip — nguoi dung khong biet da dat gioi han"),
    ("BUG-02","TC06","Khong cho phep tang thoi gian nau khi thiet bi dang hoat dong","Trung binh","Thiet bi phat tieng bip canh bao, tu choi tang thoi gian, khong hien thi thong bao giai thich"),
    ("BUG-03","TC11","Khong co chi bao truc quan khi dat nhiet do toi da 200C","Thap","Chi phat tieng bip, khong co chi bao truc quan — nguoi dung kho nhan biet da dat gioi han"),
    ("BUG-04","TC12","Khong co chi bao truc quan khi dat nhiet do toi thieu 80C","Thap","Chi phat tieng bip, khong co chi bao truc quan — nguoi dung kho nhan biet da dat gioi han"),
    ("BUG-05","TC13","Mat toan bo cai dat sau khi mat dien dot ngot","Trung binh","Thiet bi reset ve mac dinh — nguoi dung phai cai dat lai tu dau"),
]

VIDEOS = [
    ("V1","TC01","https://www.youtube.com/shorts/HfsBBp0TVtE"),
    ("V2","TC02","https://www.youtube.com/shorts/DH9anK0nrpY"),
    ("V3","TC03","https://www.youtube.com/shorts/DH9anK0nrpY"),
    ("V4","TC04","https://www.youtube.com/shorts/DH9anK0nrpY"),
    ("V5","TC07","https://www.youtube.com/shorts/bWeIbCaRzLs"),
    ("V6","TC14","https://youtube.com/shorts/x6u1C93hVnA"),
]

# SHEET 1: Test Cases
ws1 = wb.active
ws1.title = "Test Cases"
ws1.freeze_panes = "A3"
ws1.merge_cells("A1:I1")
t = ws1.cell(row=1, column=1, value="HW01 - Test Cases | Philips HD9252 Air Fryer | Nguyen Thanh Dang - 23127334")
t.fill = PatternFill("solid", fgColor=C_HEAD)
t.font = Font(color="FFFFFF", bold=True, size=12)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

COLS1  = ["TC#","Ten Test Case","Muc Tieu","Dau Vao","Cac Buoc","Ket Qua Mong Doi","Ket Qua Thuc Te","Ket Luan","Ghi Chu / Bug"]
WIDTHS1 = [10,28,38,28,50,42,42,12,22]
for i,(h,w) in enumerate(zip(COLS1,WIDTHS1),1):
    hdr(ws1,2,i,h); col_w(ws1,i,w)
ws1.row_dimensions[2].height = 28

for ri,tc in enumerate(TCS,3):
    tc_id,name,obj,inp,steps,exp,act,verdict,is_edge,note = tc
    bg = C_EDGE if is_edge else None
    v_bg = C_PASS if verdict=="Dat" else C_FAIL
    dat(ws1,ri,1,tc_id,bold=True,bg=bg,align="center",wrap=False)
    dat(ws1,ri,2,name,bold=True,bg=bg)
    dat(ws1,ri,3,obj,bg=bg)
    dat(ws1,ri,4,inp,bg=bg)
    dat(ws1,ri,5,steps,bg=bg)
    dat(ws1,ri,6,exp,bg=bg)
    dat(ws1,ri,7,act,bg=bg)
    dat(ws1,ri,8,verdict,bg=v_bg,align="center",wrap=False,bold=True)
    dat(ws1,ri,9,note,bg=bg)
    ws1.row_dimensions[ri].height = 110

lr = len(TCS)+4
ws1.cell(row=lr,column=1,value="Legend:").font = Font(bold=True,size=10)
for i,(lbl,clr) in enumerate([("Edge case AI bo sot",C_EDGE),("Dat",C_PASS),("Khong dat",C_FAIL)]):
    c = ws1.cell(row=lr+1+i,column=1,value=lbl)
    c.fill = PatternFill("solid",fgColor=clr); c.font = Font(bold=True,size=9); c.border = thin()

# SHEET 2: Execution Summary
ws2 = wb.create_sheet("Execution Summary")
ws2.freeze_panes = "A3"
ws2.merge_cells("A1:G1")
t2 = ws2.cell(row=1,column=1,value="HW01 - Tom Tat Thuc Thi | Philips HD9252 | Nguyen Thanh Dang - 23127334")
t2.fill = PatternFill("solid",fgColor=C_HEAD)
t2.font = Font(color="FFFFFF",bold=True,size=12)
t2.alignment = Alignment(horizontal="center",vertical="center")
ws2.row_dimensions[1].height = 28

COLS2  = ["TC#","Ten Test Case","Loai","Da thuc hien?","Co video?","Co loi?","Ket luan"]
WIDTHS2 = [10,42,12,14,12,12,14]
for i,(h,w) in enumerate(zip(COLS2,WIDTHS2),1):
    hdr(ws2,2,i,h); col_w(ws2,i,w)
ws2.row_dimensions[2].height = 28

for ri,(tc_id,name,typ,exe,vid,bug,verdict) in enumerate(EXEC,3):
    is_edge = "Edge" in tc_id
    bg = C_EDGE if is_edge else None
    v_bg = C_PASS if verdict=="Dat" else C_FAIL
    dat(ws2,ri,1,tc_id,bold=True,bg=bg,align="center",wrap=False)
    dat(ws2,ri,2,name,bold=is_edge,bg=bg)
    dat(ws2,ri,3,typ,bg=bg,align="center",wrap=False)
    dat(ws2,ri,4,exe,bg=bg,align="center",wrap=False)
    dat(ws2,ri,5,vid,bg=bg,align="center",wrap=False)
    dat(ws2,ri,6,bug,bg=bg,align="center",wrap=False)
    dat(ws2,ri,7,verdict,bg=v_bg,align="center",bold=True,wrap=False)
    ws2.row_dimensions[ri].height = 18

tr = len(EXEC)+4
pass_c = sum(1 for e in EXEC if e[6]=="Dat")
fail_c = sum(1 for e in EXEC if e[6]=="Khong dat")
ws2.cell(row=tr,column=1,value="Tong ket:").font = Font(bold=True)
c = ws2.cell(row=tr,column=2,value=f"Tong: {len(EXEC)} TC  |  Dat: {pass_c}  |  Khong dat: {fail_c}  |  Edge cases: 3  |  Videos: {len(VIDEOS)}")
c.font = Font(bold=True,size=10)
ws2.merge_cells(f"B{tr}:G{tr}")

# SHEET 3: Defect Report (KHONG CO COT GITHUB ISSUE)
ws3 = wb.create_sheet("Defect Report")
ws3.freeze_panes = "A3"
ws3.merge_cells("A1:E1")
t3 = ws3.cell(row=1,column=1,value="HW01 - Defect Report | Philips HD9252")
t3.fill = PatternFill("solid",fgColor=C_HEAD)
t3.font = Font(color="FFFFFF",bold=True,size=12)
t3.alignment = Alignment(horizontal="center",vertical="center")
ws3.row_dimensions[1].height = 28

COLS3  = ["Bug ID","TC lien quan","Mo ta loi","Muc do","Ket qua thuc te"]
WIDTHS3 = [10,12,45,13,55]
for i,(h,w) in enumerate(zip(COLS3,WIDTHS3),1):
    hdr(ws3,2,i,h); col_w(ws3,i,w)

SEV_BG = {"Thap":"C6EFCE","Trung binh":"FFEB9C","Cao":"FFC7CE"}
for ri,(bug_id,tc,desc,sev,actual) in enumerate(BUGS,3):
    s_bg = SEV_BG.get(sev)
    dat(ws3,ri,1,bug_id,bold=True,align="center",wrap=False)
    dat(ws3,ri,2,tc,align="center",wrap=False)
    dat(ws3,ri,3,desc)
    dat(ws3,ri,4,sev,bg=s_bg,align="center",wrap=False,bold=True)
    dat(ws3,ri,5,actual)
    ws3.row_dimensions[ri].height = 45

# SHEET 4: Test Summary Report
ws4 = wb.create_sheet("Test Summary Report")
col_w(ws4,1,30); col_w(ws4,2,55); col_w(ws4,3,18)
ws4.merge_cells("A1:C1")
t4 = ws4.cell(row=1,column=1,value="HW01 - Test Summary Report")
t4.fill = PatternFill("solid",fgColor=C_HEAD)
t4.font = Font(color="FFFFFF",bold=True,size=14)
t4.alignment = Alignment(horizontal="center",vertical="center")
ws4.row_dimensions[1].height = 32

sections = [
    ("Thong tin bai tap",[
        ("Assignment","HW01 - QA/QC Jobs - 20 Defects - Test a Physical Product"),
        ("Student ID","23127334"),("Ho ten","Nguyen Thanh Dang"),
        ("Ngay nop","[dien ngay nop]"),
    ]),
    ("Thong tin thiet bi",[
        ("Loai","Noi chien khong dau (Air Fryer)"),("Thuong hieu","Philips"),
        ("Model","HD9252 (RapidAir)"),("Nam san xuat","~2021-2023"),
        ("So se-ri","PH26****89VN"),
    ]),
    ("Pham vi kiem thu",[
        ("Tong so TC thiet ke",f"{len(TCS)}"),
        ("Edge cases (AI bo sot)","3  (TC06, TC16, TC17)"),
        ("TC da thuc thi",f"{len(EXEC)} / {len(TCS)}"),
        ("TC co video",f"{len(VIDEOS)}  (TC01, TC02, TC03, TC04, TC07, TC14)"),
    ]),
    ("Ket qua kiem thu",[
        ("Dat",f"{pass_c}  TC  ({pass_c*100//len(EXEC)}%)"),
        ("Khong dat",f"{fail_c}  TC  ({fail_c*100//len(EXEC)}%)"),
        ("Blocked","0"),("Chua chay","0"),
    ]),
    ("Defects tim duoc",[
        ("Tong so defects",f"{len(BUGS)}"),
        ("Muc Thap","3  (BUG-01, BUG-03, BUG-04)"),
        ("Muc Trung binh","2  (BUG-02, BUG-05)"),
    ]),
    ("AI Collaboration",[
        ("Cong cu AI","Claude (claude-sonnet-4-6)"),
        ("Ti le edge case AI bo sot","3 / 17  (~17.6%)"),
        ("Ket luan","AI huu ich cho TC chuc nang chuan, nhung can nguoi kiem tra phat hien edge case dua tren tuong tac thuc te voi thiet bi."),
    ]),
    ("Video YouTube (Unlisted)",[(v[0]+" - "+v[1],v[2]) for v in VIDEOS]),
]

cur = 2
for sec_title,rows in sections:
    ws4.merge_cells(f"A{cur}:C{cur}")
    sh = ws4.cell(row=cur,column=1,value=sec_title)
    sh.fill = PatternFill("solid",fgColor=C_SEC)
    sh.font = Font(color="FFFFFF",bold=True,size=11)
    sh.alignment = Alignment(horizontal="left",vertical="center")
    sh.border = thin()
    ws4.row_dimensions[cur].height = 20
    cur += 1
    for label,value in rows:
        lc = ws4.cell(row=cur,column=1,value=label)
        lc.font = Font(bold=True,size=10)
        lc.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
        lc.fill = PatternFill("solid",fgColor=C_LROW)
        lc.border = thin()
        ws4.merge_cells(f"B{cur}:C{cur}")
        vc = ws4.cell(row=cur,column=2,value=value)
        vc.font = Font(size=10)
        vc.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
        vc.border = thin()
        ws4.row_dimensions[cur].height = 18
        cur += 1
    cur += 1

out = r"C:\Users\dn156\source\software-testing\software-testing\hw\hw1\HW01_TestCases.xlsx"
wb.save(out)
print(f"Saved: {out}")
